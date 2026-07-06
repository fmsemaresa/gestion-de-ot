from fastapi import FastAPI, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from typing import List, Optional
from datetime import datetime
from sqlmodel import Session, select
from contextlib import asynccontextmanager
import io
import openpyxl

from models import (
    Planta, Edificio, Ubicacion, Activo, 
    ComponenteActivo, Tecnico, OrdenTrabajo,
    PlantillaChequeo, ItemPlantillaChequeo, RespuestaChequeo, FotoOrdenTrabajo, ComentarioAvanceOT, OcupanteUbicacion,
    OrdenTrabajoComponente, OrdenTrabajoTecnicoLink
)

from database import engine, init_db

@asynccontextmanager
async def lifespan(app: FastAPI):
    # Inicializar Base de Datos y precargar plantas/edificios/técnicos
    init_db()
    yield

from pydantic import BaseModel

class ComponenteTrabajadoInput(BaseModel):
    componente_id: int
    comentario: Optional[str] = None

class OrdenTrabajoCreate(BaseModel):
    descripcion: str
    tipo: str = "Correctiva"
    prioridad: str = "Media"
    planta_id: int
    edificio_id: int
    ubicacion_id: Optional[int] = None
    activo_id: Optional[int] = None
    tecnico_id: Optional[int] = None
    tecnico_ids: Optional[List[int]] = None
    plantilla_id: Optional[int] = None
    fecha_programada: Optional[datetime] = None
    fecha_fin_programada: Optional[datetime] = None
    reportado_por: Optional[str] = "Administración"
    componentes_trabajados: Optional[List[ComponenteTrabajadoInput]] = None
    custom_checklist_items: Optional[List[dict]] = None

app = FastAPI(
    title="Gestión de OTs - FMS Climatización API", 
    description="API para la gestión de activos, ubicaciones y órdenes de trabajo",
    lifespan=lifespan
)

# Configurar CORS por si se usa en desarrollo desacoplado
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Dependencia para obtener la sesión de base de datos
def get_db():
    with Session(engine) as session:
        yield session

# Redirección de la raíz al portal del administrador
@app.get("/")
def read_root():
    return RedirectResponse(url="/static/index.html")

# --- ENDPOINTS JERARQUÍA ---

@app.get("/api/plantas", response_model=List[Planta])
def get_plantas(db: Session = Depends(get_db)):
    return db.exec(select(Planta)).all()

@app.get("/api/plantas/{planta_id}/edificios", response_model=List[Edificio])
def get_edificios(planta_id: int, db: Session = Depends(get_db)):
    return db.exec(select(Edificio).where(Edificio.planta_id == planta_id)).all()

@app.get("/api/edificios/{edificio_id}/ubicaciones")
def get_ubicaciones(edificio_id: int, db: Session = Depends(get_db)):
    ubicaciones = db.exec(select(Ubicacion).where(Ubicacion.edificio_id == edificio_id)).all()
    return [{
        "id": u.id,
        "nombre": u.nombre,
        "codigo": u.codigo,
        "uso": u.uso,
        "cargo": u.cargo,
        "edificio_id": u.edificio_id,
        "ocupantes": [{"id": o.id, "nombre": o.nombre} for o in u.ocupantes]
    } for u in ubicaciones]

@app.get("/api/ubicaciones/{ubicacion_id}")
def get_ubicacion_by_id(ubicacion_id: int, db: Session = Depends(get_db)):
    u = db.get(Ubicacion, ubicacion_id)
    if not u:
        raise HTTPException(status_code=404, detail="Ubicación no encontrada")
    edificio = db.get(Edificio, u.edificio_id)
    planta = db.get(Planta, edificio.planta_id) if edificio else None
    return {
        "id": u.id,
        "nombre": u.nombre,
        "codigo": u.codigo,
        "uso": u.uso,
        "cargo": u.cargo,
        "edificio_id": u.edificio_id,
        "edificio_nombre": edificio.nombre if edificio else None,
        "planta_id": edificio.planta_id if edificio else None,
        "planta_nombre": planta.nombre if planta else None,
        "ocupantes": [{"id": o.id, "nombre": o.nombre} for o in u.ocupantes]
    }

@app.get("/api/search/locations")
def search_locations(db: Session = Depends(get_db)):
    results = []
    ubicaciones = db.exec(select(Ubicacion)).all()
    for u in ubicaciones:
        edificio = db.get(Edificio, u.edificio_id)
        planta = db.get(Planta, edificio.planta_id) if edificio else None
        
        # Get active assets (not replaced/deleted)
        activos = db.exec(
            select(Activo)
            .where(Activo.ubicacion_id == u.id)
            .where(Activo.estado != "Reemplazado")
            .where(Activo.estado != "Eliminado sin Reemplazo")
        ).all()
        nombres_activos = [a.nombre for a in activos]
        
        # Get occupants names
        ocupantes_list = [{"id": o.id, "nombre": o.nombre} for o in u.ocupantes]
        
        results.append({
            "id": u.id,
            "nombre": u.nombre,
            "codigo": u.codigo,
            "uso": u.uso,
            "cargo": u.cargo,
            "edificio_id": u.edificio_id,
            "edificio_nombre": edificio.nombre if edificio else "N/A",
            "planta_id": edificio.planta_id if edificio else None,
            "planta_nombre": planta.nombre if planta else "N/A",
            "activos": nombres_activos,
            "ocupantes": ocupantes_list
        })
    return results

@app.post("/api/ubicaciones", response_model=Ubicacion)
def create_ubicacion(ubicacion: Ubicacion, db: Session = Depends(get_db)):
    # Verificar que el edificio existe
    edificio = db.get(Edificio, ubicacion.edificio_id)
    if not edificio:
        raise HTTPException(status_code=404, detail="Edificio no encontrado")
    
    db.add(ubicacion)
    db.commit()
    db.refresh(ubicacion)
    return ubicacion

@app.put("/api/ubicaciones/{ubicacion_id}", response_model=Ubicacion)
def update_ubicacion(ubicacion_id: int, updated_ubic: Ubicacion, db: Session = Depends(get_db)):
    ubic = db.get(Ubicacion, ubicacion_id)
    if not ubic:
        raise HTTPException(status_code=404, detail="Ubicación no encontrada")
        
    ubic.nombre = updated_ubic.nombre
    if updated_ubic.codigo:
        ubic.codigo = updated_ubic.codigo
    if updated_ubic.uso:
        ubic.uso = updated_ubic.uso
    if updated_ubic.cargo:
        ubic.cargo = updated_ubic.cargo
    if updated_ubic.color:
        ubic.color = updated_ubic.color
    if updated_ubic.edificio_id:
        ubic.edificio_id = updated_ubic.edificio_id
        
    db.add(ubic)
    db.commit()
    db.refresh(ubic)
    return ubic

# --- ENDPOINTS ACTIVOS Y DESPIECE ---

@app.get("/api/activos")
def get_activos(
    planta_id: Optional[int] = None,
    edificio_id: Optional[int] = None,
    ubicacion_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    query = select(Activo)
    
    if ubicacion_id:
        query = query.where(Activo.ubicacion_id == ubicacion_id)
    elif edificio_id:
        # Unir con Ubicacion para filtrar por edificio
        query = query.join(Ubicacion).where(Ubicacion.edificio_id == edificio_id)
    elif planta_id:
        # Unir con Ubicacion y Edificio para filtrar por planta
        query = query.join(Ubicacion).join(Edificio).where(Edificio.planta_id == planta_id)
        
    activos = db.exec(query).all()
    
    results = []
    for a in activos:
        ubicacion = db.get(Ubicacion, a.ubicacion_id) if a.ubicacion_id else None
        edificio = db.get(Edificio, ubicacion.edificio_id) if ubicacion else None
        planta = db.get(Planta, edificio.planta_id) if edificio else None
        
        # Check if there is an associated PlantillaChequeo for this asset
        plantilla = db.exec(select(PlantillaChequeo).where(PlantillaChequeo.activo_id == a.id)).first()
        
        results.append({
            "id": a.id,
            "nombre": a.nombre,
            "tipo": a.tipo,
            "marca": a.marca,
            "modelo": a.modelo,
            "numero_serie": a.numero_serie,
            "estado": a.estado,
            "color": a.color,
            "ubicacion_id": a.ubicacion_id,
            "ubicacion_nombre": ubicacion.nombre if ubicacion else None,
            "edificio_nombre": edificio.nombre if edificio else None,
            "planta_nombre": planta.nombre if planta else None,
            "cantidad_despiece": len(a.componentes) if a.componentes else 0,
            "plantilla_id": plantilla.id if plantilla else None
        })
        
    return results

@app.get("/api/activos/{activo_id}")
def get_activo_detail(activo_id: int, db: Session = Depends(get_db)):
    activo = db.get(Activo, activo_id)
    if not activo:
        raise HTTPException(status_code=404, detail="Activo no encontrado")
    
    # Obtener componentes (despiece)
    componentes = db.exec(select(ComponenteActivo).where(ComponenteActivo.activo_id == activo_id)).all()
    # Obtener historial de OTs
    ots = db.exec(select(OrdenTrabajo).where(OrdenTrabajo.activo_id == activo_id)).all()
    
    # Obtener información de ubicación jerárquica para la ficha
    ubicacion = db.get(Ubicacion, activo.ubicacion_id) if activo.ubicacion_id else None
    edificio = db.get(Edificio, ubicacion.edificio_id) if ubicacion else None
    planta = db.get(Planta, edificio.planta_id) if edificio else None
    
    # Check if there is an associated PlantillaChequeo for this asset
    plantilla = db.exec(select(PlantillaChequeo).where(PlantillaChequeo.activo_id == activo_id)).first()
    
    return {
        "id": activo.id,
        "nombre": activo.nombre,
        "tipo": activo.tipo,
        "marca": activo.marca,
        "modelo": activo.modelo,
        "numero_serie": activo.numero_serie,
        "estado": activo.estado,
        "ubicacion_id": activo.ubicacion_id,
        "ubicacion_nombre": ubicacion.nombre if ubicacion else None,
        "edificio_id": edificio.id if edificio else None,
        "edificio_nombre": edificio.nombre if edificio else None,
        "planta_id": planta.id if planta else None,
        "planta_nombre": planta.nombre if planta else None,
        "componentes": componentes,
        "ordenes_trabajo": ots,
        "plantilla_id": plantilla.id if plantilla else None
    }

@app.post("/api/activos", response_model=Activo)
def create_activo(activo: Activo, db: Session = Depends(get_db)):
    if activo.ubicacion_id:
        ubicacion = db.get(Ubicacion, activo.ubicacion_id)
        if not ubicacion:
            raise HTTPException(status_code=404, detail="Ubicación no encontrada")
            
    db.add(activo)
    db.commit()
    db.refresh(activo)
    return activo

@app.post("/api/componentes", response_model=ComponenteActivo)
def create_componente(componente: ComponenteActivo, db: Session = Depends(get_db)):
    activo = db.get(Activo, componente.activo_id)
    if not activo:
        raise HTTPException(status_code=404, detail="Activo principal no encontrado")
        
    db.add(componente)
    db.commit()
    db.refresh(componente)
    return componente

@app.put("/api/activos/{activo_id}", response_model=Activo)
def update_activo(activo_id: int, updated_activo: Activo, db: Session = Depends(get_db)):
    activo = db.get(Activo, activo_id)
    if not activo:
        raise HTTPException(status_code=404, detail="Activo no encontrado")
        
    activo.nombre = updated_activo.nombre
    activo.tipo = updated_activo.tipo
    activo.marca = updated_activo.marca
    activo.modelo = updated_activo.modelo
    activo.numero_serie = updated_activo.numero_serie
    activo.estado = updated_activo.estado
    if updated_activo.ubicacion_id:
        activo.ubicacion_id = updated_activo.ubicacion_id
        
    db.add(activo)
    db.commit()
    db.refresh(activo)
    return activo

@app.put("/api/componentes/{componente_id}", response_model=ComponenteActivo)
def update_componente(componente_id: int, updated_comp: ComponenteActivo, db: Session = Depends(get_db)):
    comp = db.get(ComponenteActivo, componente_id)
    if not comp:
        raise HTTPException(status_code=404, detail="Componente no encontrado")
        
    comp.nombre = updated_comp.nombre
    comp.marca = updated_comp.marca
    comp.modelo = updated_comp.modelo
    comp.numero_serie = updated_comp.numero_serie
    comp.estado = updated_comp.estado
    comp.activo_id = updated_comp.activo_id
    
    db.add(comp)
    db.commit()
    db.refresh(comp)
    return comp

# --- ENDPOINTS TÉCNICOS ---

@app.get("/api/tecnicos", response_model=List[Tecnico])
def get_tecnicos(db: Session = Depends(get_db)):
    return db.exec(select(Tecnico).order_by(Tecnico.id)).all()

# --- ENDPOINTS ÓRDENES DE TRABAJO (OT) ---

@app.get("/api/ordenes")
def get_ordenes(
    estado: Optional[str] = None,
    tecnico_id: Optional[int] = None,
    planta_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    query = select(OrdenTrabajo)
    
    if tecnico_id:
        query = query.join(OrdenTrabajoTecnicoLink, OrdenTrabajoTecnicoLink.orden_trabajo_id == OrdenTrabajo.id).where(OrdenTrabajoTecnicoLink.tecnico_id == tecnico_id)
    if planta_id:
        query = query.where(OrdenTrabajo.planta_id == planta_id)
        
    # Ordenar por ID descendente para ver las más nuevas primero
    query = query.order_by(OrdenTrabajo.id.desc())
    
    ots = db.exec(query).all()
    
    # Enriquecer respuesta con nombres legibles
    results = []
    for ot in ots:
        # Mapeo dinámico del estado
        mapped_estado = ot.estado
        if mapped_estado in ("Resuelta", "REALIZADA"):
            mapped_estado = "REALIZADA"
        elif mapped_estado != "Cancelada":
            # Si tiene técnicos asignados, evaluar estado
            if ot.tecnicos:
                if ot.fecha_programada:
                    mapped_estado = "PROGRAMADA"
                else:
                    mapped_estado = "ASIGNADA"
            else:
                mapped_estado = "CREADA"
        
        # Filtrar por estado si es requerido
        if estado:
            target_states = [estado]
            if estado == "REALIZADA":
                target_states.append("Resuelta")
            elif estado == "Resuelta":
                target_states.append("REALIZADA")
            
            if mapped_estado not in target_states:
                continue
                
        planta = db.get(Planta, ot.planta_id)
        edificio = db.get(Edificio, ot.edificio_id)
        ubicacion = db.get(Ubicacion, ot.ubicacion_id) if ot.ubicacion_id else None
        activo = db.get(Activo, ot.activo_id) if ot.activo_id else None
        
        tecnicos_list = ot.tecnicos
        tecnico_nombre = ", ".join([t.nombre for t in tecnicos_list]) if tecnicos_list else "No asignado"
        tecnico_ids = [t.id for t in tecnicos_list]
        
        # Obtener fotos
        fotos_db = db.exec(select(FotoOrdenTrabajo).where(FotoOrdenTrabajo.orden_trabajo_id == ot.id)).all()
        fotos_list = [{
            "id": f.id,
            "url_foto": f.url_foto,
            "comentario": f.comentario,
            "fecha_creacion": f.fecha_creacion.isoformat() + "Z" if f.fecha_creacion else None
        } for f in fotos_db]
        
        # Obtener comentarios de avance (ordenados por fecha ascendente para ver el hilo de conversación)
        comentarios_db = db.exec(
            select(ComentarioAvanceOT)
            .where(ComentarioAvanceOT.orden_trabajo_id == ot.id)
            .order_by(ComentarioAvanceOT.fecha_creacion.asc())
        ).all()
        comentarios_list = [{
            "id": c.id,
            "comentario": c.comentario,
            "autor": c.autor,
            "fecha_creacion": c.fecha_creacion.isoformat() + "Z" if c.fecha_creacion else None
        } for c in comentarios_db]
        
        # Obtener componentes seleccionados y sus comentarios individuales
        componentes_db = db.exec(
            select(OrdenTrabajoComponente, ComponenteActivo)
            .join(ComponenteActivo, OrdenTrabajoComponente.componente_id == ComponenteActivo.id)
            .where(OrdenTrabajoComponente.orden_trabajo_id == ot.id)
        ).all()
        componentes_list = [{
            "id": comp.id,
            "nombre": comp.nombre,
            "comentario": ot_comp.comentario
        } for ot_comp, comp in componentes_db]
        
        results.append({
            "id": ot.id,
            "descripcion": ot.descripcion,
            "tipo": ot.tipo,
            "estado": mapped_estado,
            "prioridad": ot.prioridad,
            "fecha_creacion": ot.fecha_creacion.isoformat() + "Z" if ot.fecha_creacion else None,
            "fecha_resolucion": ot.fecha_resolucion.isoformat() + "Z" if ot.fecha_resolucion else None,
            "fecha_programada": ot.fecha_programada.isoformat() if ot.fecha_programada else None,
            "fecha_fin_programada": ot.fecha_fin_programada.isoformat() if ot.fecha_fin_programada else None,
            "fecha_inicio": ot.fecha_inicio.isoformat() + "Z" if ot.fecha_inicio else None,
            "estado_ejecucion": ot.estado_ejecucion,
            "reportado_por": ot.reportado_por,
            "comentarios_tecnicos": ot.comentarios_tecnicos,
            "planta_id": ot.planta_id,
            "edificio_id": ot.edificio_id,
            "ubicacion_id": ot.ubicacion_id,
            "activo_id": ot.activo_id,
            "planta_nombre": planta.nombre if planta else None,
            "edificio_nombre": edificio.nombre if edificio else None,
            "ubicacion_nombre": ubicacion.nombre if ubicacion else None,
            "activo_nombre": activo.nombre if activo else None,
            "activo_tipo": activo.tipo if activo else "Otros",
            "activo_color": activo.color if activo else None,
            "ubicacion_color": ubicacion.color if ubicacion else None,
            "tecnico_nombre": tecnico_nombre,
            "tecnico_id": ot.tecnico_id,
            "tecnico_ids": tecnico_ids,
            "tecnicos": [{"id": t.id, "nombre": t.nombre} for t in tecnicos_list],
            "plantilla_id": ot.plantilla_id,
            "fotos": fotos_list,
            "comentarios_avance": comentarios_list,
            "componentes_trabajados": componentes_list
        })
        
    return results

@app.get("/api/ordenes/{ot_id}")
def get_orden_by_id(ot_id: int, db: Session = Depends(get_db)):
    ot = db.get(OrdenTrabajo, ot_id)
    if not ot:
        raise HTTPException(status_code=404, detail="Orden de trabajo no encontrada")
        
    # Mapeo dinámico del estado
    mapped_estado = ot.estado
    if mapped_estado in ("Resuelta", "REALIZADA"):
        mapped_estado = "REALIZADA"
    elif mapped_estado != "Cancelada":
        if ot.tecnicos:
            if ot.fecha_programada:
                mapped_estado = "PROGRAMADA"
            else:
                mapped_estado = "ASIGNADA"
        else:
            mapped_estado = "CREADA"
            
    planta = db.get(Planta, ot.planta_id)
    edificio = db.get(Edificio, ot.edificio_id)
    ubicacion = db.get(Ubicacion, ot.ubicacion_id) if ot.ubicacion_id else None
    activo = db.get(Activo, ot.activo_id) if ot.activo_id else None
    
    tecnicos_list = ot.tecnicos
    tecnico_nombre = ", ".join([t.nombre for t in tecnicos_list]) if tecnicos_list else "No asignado"
    tecnico_ids = [t.id for t in tecnicos_list]
    
    # Obtener fotos
    fotos_db = db.exec(select(FotoOrdenTrabajo).where(FotoOrdenTrabajo.orden_trabajo_id == ot.id)).all()
    fotos_list = [{
        "id": f.id,
        "url_foto": f.url_foto,
        "comentario": f.comentario,
        "fecha_creacion": f.fecha_creacion.isoformat() + "Z" if f.fecha_creacion else None
    } for f in fotos_db]
    
    # Obtener comentarios de avance
    comentarios_db = db.exec(
        select(ComentarioAvanceOT)
        .where(ComentarioAvanceOT.orden_trabajo_id == ot.id)
        .order_by(ComentarioAvanceOT.fecha_creacion.asc())
    ).all()
    comentarios_list = [{
        "id": c.id,
        "comentario": c.comentario,
        "autor": c.autor,
        "fecha_creacion": c.fecha_creacion.isoformat() + "Z" if c.fecha_creacion else None
    } for c in comentarios_db]
    
    # Obtener componentes seleccionados y sus comentarios individuales
    componentes_db = db.exec(
        select(OrdenTrabajoComponente, ComponenteActivo)
        .join(ComponenteActivo, OrdenTrabajoComponente.componente_id == ComponenteActivo.id)
        .where(OrdenTrabajoComponente.orden_trabajo_id == ot.id)
    ).all()
    componentes_list = [{
        "id": comp.id,
        "nombre": comp.nombre,
        "comentario": ot_comp.comentario
    } for ot_comp, comp in componentes_db]
    
    return {
        "id": ot.id,
        "descripcion": ot.descripcion,
        "tipo": ot.tipo,
        "estado": mapped_estado,
        "prioridad": ot.prioridad,
        "fecha_creacion": ot.fecha_creacion.isoformat() + "Z" if ot.fecha_creacion else None,
        "fecha_resolucion": ot.fecha_resolucion.isoformat() + "Z" if ot.fecha_resolucion else None,
        "fecha_programada": ot.fecha_programada.isoformat() if ot.fecha_programada else None,
        "fecha_fin_programada": ot.fecha_fin_programada.isoformat() if ot.fecha_fin_programada else None,
        "fecha_inicio": ot.fecha_inicio.isoformat() + "Z" if ot.fecha_inicio else None,
        "estado_ejecucion": ot.estado_ejecucion,
        "reportado_por": ot.reportado_por,
        "comentarios_tecnicos": ot.comentarios_tecnicos,
        "planta_id": ot.planta_id,
        "edificio_id": ot.edificio_id,
        "ubicacion_id": ot.ubicacion_id,
        "activo_id": ot.activo_id,
        "planta_nombre": planta.nombre if planta else None,
        "edificio_nombre": edificio.nombre if edificio else None,
        "ubicacion_nombre": ubicacion.nombre if ubicacion else None,
        "activo_nombre": activo.nombre if activo else None,
        "activo_tipo": activo.tipo if activo else "Otros",
        "activo_color": activo.color if activo else None,
        "ubicacion_color": ubicacion.color if ubicacion else None,
        "tecnico_nombre": tecnico_nombre,
        "tecnico_id": ot.tecnico_id,
        "tecnico_ids": tecnico_ids,
        "tecnicos": [{"id": t.id, "nombre": t.nombre} for t in tecnicos_list],
        "plantilla_id": ot.plantilla_id,
        "fotos": fotos_list,
        "comentarios_avance": comentarios_list,
        "componentes_trabajados": componentes_list
    }

@app.post("/api/ordenes", response_model=OrdenTrabajo)
def create_orden(ot_in: OrdenTrabajoCreate, db: Session = Depends(get_db)):
    # Validar que los campos de ubicación sean correctos
    planta = db.get(Planta, ot_in.planta_id)
    edificio = db.get(Edificio, ot_in.edificio_id)
    if not planta or not edificio:
        raise HTTPException(status_code=404, detail="Planta o Edificio no válido")
        
    if edificio.nombre == "NZ":
        # Ubicación es opcional para el sector especial NZ
        pass
    else:
        # Ubicación es obligatoria para cualquier otro edificio
        if not ot_in.ubicacion_id:
            raise HTTPException(status_code=400, detail="La ubicación es obligatoria para este edificio.")
            
    if ot_in.ubicacion_id:
        ubicacion = db.get(Ubicacion, ot_in.ubicacion_id)
        if not ubicacion:
            raise HTTPException(status_code=404, detail="Ubicación no encontrada")
            
    if ot_in.activo_id:
        activo = db.get(Activo, ot_in.activo_id)
        if not activo:
            raise HTTPException(status_code=404, detail="Activo no encontrado")
            
    # Calcular automáticamente el estado
    estado = "CREADA"
    assigned_tecnicos = []
    primary_tecnico_id = None
    
    if ot_in.tecnico_ids is not None:
        for tid in ot_in.tecnico_ids:
            tech = db.get(Tecnico, tid)
            if tech:
                assigned_tecnicos.append(tech)
        if assigned_tecnicos:
            primary_tecnico_id = assigned_tecnicos[0].id
    elif ot_in.tecnico_id:
        tech = db.get(Tecnico, ot_in.tecnico_id)
        if tech:
            assigned_tecnicos.append(tech)
            primary_tecnico_id = tech.id
            
    if assigned_tecnicos:
        if ot_in.fecha_programada:
            estado = "PROGRAMADA"
        else:
            estado = "ASIGNADA"
    else:
        estado = "CREADA"
        
    if ot_in.fecha_programada and ot_in.fecha_fin_programada:
        if ot_in.fecha_fin_programada < ot_in.fecha_programada:
            raise HTTPException(status_code=400, detail="La fecha/hora de término no puede ser anterior a la de inicio.")

    ot = OrdenTrabajo(
        descripcion=ot_in.descripcion,
        tipo=ot_in.tipo,
        estado=estado,
        prioridad=ot_in.prioridad,
        fecha_programada=ot_in.fecha_programada,
        fecha_fin_programada=ot_in.fecha_fin_programada,
        reportado_por=ot_in.reportado_por,
        planta_id=ot_in.planta_id,
        edificio_id=ot_in.edificio_id,
        ubicacion_id=ot_in.ubicacion_id,
        activo_id=ot_in.activo_id,
        tecnico_id=primary_tecnico_id,
        plantilla_id=ot_in.plantilla_id,
        tecnicos=assigned_tecnicos
    )
    
    db.add(ot)
    db.commit()
    db.refresh(ot)
    
    # Guardar componentes seleccionados
    if ot_in.componentes_trabajados:
        for comp_in in ot_in.componentes_trabajados:
            link = OrdenTrabajoComponente(
                orden_trabajo_id=ot.id,
                componente_id=comp_in.componente_id,
                comentario=comp_in.comentario
            )
            db.add(link)
        db.commit()
        db.refresh(ot)
        
    # Guardar items de checklist personalizados si existen
    if ot_in.custom_checklist_items:
        # 1. Crear plantilla personalizada para esta OT
        custom_plantilla = PlantillaChequeo(
            nombre=f"Chequeo Personalizado OT #{ot.id}",
            descripcion=f"Plantilla creada para el chequeo de la OT #{ot.id}",
            tipo_revision="Chequeo Preventivo"
        )
        db.add(custom_plantilla)
        db.commit()
        db.refresh(custom_plantilla)
        
        # 2. Si ya había seleccionado una plantilla base, copiar sus preguntas primero
        if ot_in.plantilla_id:
            base_items = db.exec(select(ItemPlantillaChequeo).where(ItemPlantillaChequeo.plantilla_id == ot_in.plantilla_id)).all()
            for bi in base_items:
                copied_item = ItemPlantillaChequeo(
                    texto_pregunta=bi.texto_pregunta,
                    tipo_respuesta=bi.tipo_respuesta,
                    unidad_medida=bi.unidad_medida,
                    plantilla_id=custom_plantilla.id
                )
                db.add(copied_item)
                
        # 3. Guardar las preguntas adicionales personalizadas
        for item_in in ot_in.custom_checklist_items:
            new_item = ItemPlantillaChequeo(
                texto_pregunta=item_in.get("texto_pregunta"),
                tipo_respuesta=item_in.get("tipo_respuesta", "booleano"),
                unidad_medida=item_in.get("unidad_medida") or None,
                plantilla_id=custom_plantilla.id
            )
            db.add(new_item)
            
        db.commit()
        
        # 4. Asignar la plantilla personalizada a la OT
        ot.plantilla_id = custom_plantilla.id
        db.add(ot)
        db.commit()
        db.refresh(ot)
        
    return ot

@app.put("/api/ordenes/{ot_id}", response_model=OrdenTrabajo)
def update_orden(ot_id: int, updated_ot: dict, db: Session = Depends(get_db)):
    ot = db.get(OrdenTrabajo, ot_id)
    if not ot:
        raise HTTPException(status_code=404, detail="Orden de Trabajo no encontrada")
        
    # Process basic creation fields if updated
    if "planta_id" in updated_ot:
        val = updated_ot["planta_id"]
        ot.planta_id = int(val) if val is not None else ot.planta_id
    if "edificio_id" in updated_ot:
        val = updated_ot["edificio_id"]
        ot.edificio_id = int(val) if val is not None else ot.edificio_id
    if "ubicacion_id" in updated_ot:
        val = updated_ot["ubicacion_id"]
        ot.ubicacion_id = int(val) if (val is not None and val != "" and str(val).isdigit()) else None
    if "activo_id" in updated_ot:
        val = updated_ot["activo_id"]
        ot.activo_id = int(val) if (val is not None and val != "" and str(val).isdigit()) else None
    if "tipo" in updated_ot:
        ot.tipo = updated_ot["tipo"]
    if "descripcion" in updated_ot:
        ot.descripcion = updated_ot["descripcion"]
    if "plantilla_id" in updated_ot:
        val = updated_ot["plantilla_id"]
        ot.plantilla_id = int(val) if (val is not None and val != "" and str(val).isdigit()) else None
    if "reportado_por" in updated_ot:
        ot.reportado_por = updated_ot["reportado_por"]

    # Validate location requirement based on final building
    edificio = db.get(Edificio, ot.edificio_id)
    if not edificio:
        raise HTTPException(status_code=404, detail="Edificio no encontrado")
    if edificio.nombre != "NZ" and not ot.ubicacion_id:
        raise HTTPException(status_code=400, detail="La ubicación es obligatoria para este edificio.")

    # Process technician assignment
    if "tecnico_ids" in updated_ot:
        tids = updated_ot["tecnico_ids"] or []
        assigned = []
        for tid in tids:
            tech = db.get(Tecnico, tid)
            if tech:
                assigned.append(tech)
        ot.tecnicos = assigned
        ot.tecnico_id = assigned[0].id if assigned else None
    elif "tecnico_id" in updated_ot:
        val = updated_ot["tecnico_id"]
        tid = int(val) if val is not None and str(val).isdigit() else None
        ot.tecnico_id = tid
        if tid:
            tech = db.get(Tecnico, tid)
            ot.tecnicos = [tech] if tech else []
        else:
            ot.tecnicos = []
        
    # Process scheduled date
    if "fecha_programada" in updated_ot:
        val = updated_ot["fecha_programada"]
        if val:
            if isinstance(val, str):
                try:
                    if "T" in val:
                        ot.fecha_programada = datetime.fromisoformat(val.replace("Z", "+00:00"))
                    else:
                        ot.fecha_programada = datetime.strptime(val, "%Y-%m-%d")
                except Exception:
                    ot.fecha_programada = datetime.fromisoformat(val)
            else:
                ot.fecha_programada = val
        else:
            ot.fecha_programada = None

    # Process scheduled end date
    if "fecha_fin_programada" in updated_ot:
        val = updated_ot["fecha_fin_programada"]
        if val:
            if isinstance(val, str):
                try:
                    if "T" in val:
                        ot.fecha_fin_programada = datetime.fromisoformat(val.replace("Z", "+00:00"))
                    else:
                        ot.fecha_fin_programada = datetime.strptime(val, "%Y-%m-%d")
                except Exception:
                    ot.fecha_fin_programada = datetime.fromisoformat(val)
            else:
                ot.fecha_fin_programada = val
        else:
            ot.fecha_fin_programada = None

    # Validate: fecha_fin_programada cannot be anterior to fecha_programada
    if ot.fecha_programada and ot.fecha_fin_programada:
        p_dt = ot.fecha_programada.replace(tzinfo=None)
        f_dt = ot.fecha_fin_programada.replace(tzinfo=None)
        if f_dt < p_dt:
            raise HTTPException(status_code=400, detail="La fecha/hora de término no puede ser anterior a la de inicio.")

    # Process start date
    if "fecha_inicio" in updated_ot:
        val = updated_ot["fecha_inicio"]
        if val:
            if isinstance(val, str):
                try:
                    if "T" in val:
                        ot.fecha_inicio = datetime.fromisoformat(val.replace("Z", "+00:00"))
                    else:
                        ot.fecha_inicio = datetime.strptime(val, "%Y-%m-%d")
                except Exception:
                    ot.fecha_inicio = datetime.fromisoformat(val)
            else:
                ot.fecha_inicio = val
            if "estado_ejecucion" not in updated_ot:
                ot.estado_ejecucion = "EN_PROCESO"
        else:
            ot.fecha_inicio = None
            if "estado_ejecucion" not in updated_ot:
                ot.estado_ejecucion = "NO_INICIADA"

    # Validate: fecha_programada cannot be posterior to fecha_inicio (when rescheduling)
    if "fecha_programada" in updated_ot and ot.fecha_programada and ot.fecha_inicio:
        p_dt = ot.fecha_programada.replace(tzinfo=None)
        
        from zoneinfo import ZoneInfo
        from datetime import timezone
        
        i_dt = ot.fecha_inicio
        if i_dt.tzinfo is None:
            i_dt = i_dt.replace(tzinfo=timezone.utc)
            
        i_local = i_dt.astimezone(ZoneInfo("America/Santiago")).replace(tzinfo=None)
        
        if p_dt > i_local:
            raise HTTPException(
                status_code=400,
                detail="La fecha programada no puede ser posterior a la fecha y hora de inicio de los trabajos."
            )

    if "prioridad" in updated_ot:
        ot.prioridad = updated_ot["prioridad"]
    if "comentarios_tecnicos" in updated_ot:
        ot.comentarios_tecnicos = updated_ot["comentarios_tecnicos"]
        
    # Process status transition
    if "estado" in updated_ot:
        req_estado = updated_ot["estado"]
        if req_estado in ("REALIZADA", "Resuelta"):
            ot.estado = "REALIZADA"
            ot.fecha_resolucion = datetime.utcnow()
            ot.estado_ejecucion = "REALIZADA"
            if ot.activo_id:
                activo = db.get(Activo, ot.activo_id)
                if activo:
                    activo.estado = "Operativo"
                    db.add(activo)
        elif req_estado == "Cancelada":
            ot.estado = "Cancelada"
        else:
            ot.estado = req_estado
            # If manually moved back to CREADA/ASIGNADA/PROGRAMADA, adjust execution status if needed
            if req_estado in ("CREADA", "ASIGNADA", "PROGRAMADA"):
                if ot.estado_ejecucion == "REALIZADA":
                    ot.estado_ejecucion = "NO_INICIADA"
                    ot.fecha_resolucion = None

    if "estado_ejecucion" in updated_ot:
        val_ej = updated_ot["estado_ejecucion"]
        ot.estado_ejecucion = val_ej
        if val_ej == "REALIZADA":
            ot.estado = "REALIZADA"
            if not ot.fecha_resolucion:
                ot.fecha_resolucion = datetime.utcnow()
            if ot.activo_id:
                activo = db.get(Activo, ot.activo_id)
                if activo:
                    activo.estado = "Operativo"
                    db.add(activo)
        else:
            # Re-evaluate non-terminal state
            if ot.tecnicos:
                if ot.fecha_programada:
                    ot.estado = "PROGRAMADA"
                else:
                    ot.estado = "ASIGNADA"
            else:
                ot.estado = "CREADA"
    else:
        # Re-evaluate state if not terminal
        if ot.estado not in ("REALIZADA", "Resuelta", "Cancelada"):
            if ot.tecnicos:
                if ot.fecha_programada:
                    ot.estado = "PROGRAMADA"
                else:
                    ot.estado = "ASIGNADA"
            else:
                ot.estado = "CREADA"
                
    if "respuestas" in updated_ot:
        respuestas_data = updated_ot["respuestas"]
        # Limpiar respuestas previas si existen
        existing_resps = db.exec(select(RespuestaChequeo).where(RespuestaChequeo.orden_trabajo_id == ot_id)).all()
        for er in existing_resps:
            db.delete(er)
        db.commit()
        
        # Guardar nuevas respuestas
        for resp in respuestas_data:
            item_id = resp.get("item_plantilla_id")
            val_bool = resp.get("valor_booleano")
            val_text = resp.get("valor_texto")
            val_num = resp.get("valor_numerico")
            obs = resp.get("observacion")
            
            if item_id is not None:
                new_resp = RespuestaChequeo(
                    orden_trabajo_id=ot_id,
                    item_plantilla_id=item_id,
                    valor_booleano=val_bool,
                    valor_texto=val_text,
                    valor_numerico=val_num,
                    observacion=obs
                )
                db.add(new_resp)
        db.commit()

    if "componentes_trabajados" in updated_ot:
        # Delete existing OrdenTrabajoComponente associations
        existing_links = db.exec(select(OrdenTrabajoComponente).where(OrdenTrabajoComponente.orden_trabajo_id == ot_id)).all()
        for link in existing_links:
            db.delete(link)
        db.commit()
        
        # Save new ones
        comps_in = updated_ot["componentes_trabajados"]
        if comps_in:
            for comp_in in comps_in:
                comp_id = comp_in.get("componente_id")
                comp_comentario = comp_in.get("comentario")
                if comp_id is not None:
                    link = OrdenTrabajoComponente(
                        orden_trabajo_id=ot_id,
                        componente_id=int(comp_id),
                        comentario=comp_comentario
                    )
                    db.add(link)
            db.commit()
        
    if "custom_checklist_items" in updated_ot:
        custom_items = updated_ot["custom_checklist_items"]
        if custom_items:
            plantilla_actual = db.get(PlantillaChequeo, ot.plantilla_id) if ot.plantilla_id else None
            
            if plantilla_actual and plantilla_actual.nombre == f"Chequeo Personalizado OT #{ot.id}":
                for item_in in custom_items:
                    exists = db.exec(
                        select(ItemPlantillaChequeo)
                        .where(ItemPlantillaChequeo.plantilla_id == plantilla_actual.id)
                        .where(ItemPlantillaChequeo.texto_pregunta == item_in.get("texto_pregunta"))
                    ).first()
                    if not exists:
                        new_item = ItemPlantillaChequeo(
                            texto_pregunta=item_in.get("texto_pregunta"),
                            tipo_respuesta=item_in.get("tipo_respuesta", "booleano"),
                            unidad_medida=item_in.get("unidad_medida") or None,
                            plantilla_id=plantilla_actual.id
                        )
                        db.add(new_item)
                db.commit()
            else:
                custom_plantilla = PlantillaChequeo(
                    nombre=f"Chequeo Personalizado OT #{ot.id}",
                    descripcion=f"Plantilla creada para el chequeo de la OT #{ot.id}",
                    tipo_revision="Chequeo Preventivo"
                )
                db.add(custom_plantilla)
                db.commit()
                db.refresh(custom_plantilla)
                
                if ot.plantilla_id:
                    base_items = db.exec(select(ItemPlantillaChequeo).where(ItemPlantillaChequeo.plantilla_id == ot.plantilla_id)).all()
                    for bi in base_items:
                        copied_item = ItemPlantillaChequeo(
                            texto_pregunta=bi.texto_pregunta,
                            tipo_respuesta=bi.tipo_respuesta,
                            unidad_medida=bi.unidad_medida,
                            plantilla_id=custom_plantilla.id
                        )
                        db.add(copied_item)
                
                for item_in in custom_items:
                    new_item = ItemPlantillaChequeo(
                        texto_pregunta=item_in.get("texto_pregunta"),
                        tipo_respuesta=item_in.get("tipo_respuesta", "booleano"),
                        unidad_medida=item_in.get("unidad_medida") or None,
                        plantilla_id=custom_plantilla.id
                    )
                    db.add(new_item)
                    
                db.commit()
                ot.plantilla_id = custom_plantilla.id
                db.add(ot)
                db.commit()
                db.refresh(ot)
                
    db.add(ot)
    db.commit()
    db.refresh(ot)
    return ot

@app.post("/api/ordenes/{ot_id}/fotos")
async def upload_ot_foto(
    ot_id: int,
    file: UploadFile = File(...),
    comentario: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    import os
    import uuid
    
    ot = db.get(OrdenTrabajo, ot_id)
    if not ot:
        raise HTTPException(status_code=404, detail="Orden de Trabajo no encontrada")
        
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in [".jpg", ".jpeg", ".png", ".gif", ".webp", ".heic"]:
        raise HTTPException(status_code=400, detail="Formato de archivo no permitido. Sube una imagen (.jpg, .png, etc.)")
        
    # Crear directorio si no existe
    uploads_dir = os.path.join("static", "uploads")
    if not os.path.exists(uploads_dir):
        os.makedirs(uploads_dir)
        
    unique_filename = f"ot_{ot_id}_{uuid.uuid4().hex}{ext}"
    file_path = os.path.join(uploads_dir, unique_filename)
    
    try:
        contents = await file.read()
        with open(file_path, "wb") as f:
            f.write(contents)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al guardar el archivo: {str(e)}")
        
    # Guardar en BD
    new_foto = FotoOrdenTrabajo(
        orden_trabajo_id=ot_id,
        url_foto=f"/static/uploads/{unique_filename}",
        comentario=comentario,
        fecha_creacion=datetime.utcnow()
    )
    db.add(new_foto)
    db.commit()
    db.refresh(new_foto)
    
    return {
        "id": new_foto.id,
        "orden_trabajo_id": new_foto.orden_trabajo_id,
        "url_foto": new_foto.url_foto,
        "comentario": new_foto.comentario,
        "fecha_creacion": new_foto.fecha_creacion.isoformat() + "Z"
    }

@app.post("/api/ordenes/{ot_id}/comentarios")
def add_ot_comentario(
    ot_id: int,
    payload: dict,
    db: Session = Depends(get_db)
):
    ot = db.get(OrdenTrabajo, ot_id)
    if not ot:
        raise HTTPException(status_code=404, detail="Orden de Trabajo no encontrada")
        
    comentario_txt = payload.get("comentario")
    autor_txt = payload.get("autor", "Sistema")
    
    if not comentario_txt:
        raise HTTPException(status_code=400, detail="El comentario no puede estar vacío")
        
    new_comentario = ComentarioAvanceOT(
        orden_trabajo_id=ot_id,
        comentario=comentario_txt,
        autor=autor_txt,
        fecha_creacion=datetime.utcnow()
    )
    db.add(new_comentario)
    db.commit()
    db.refresh(new_comentario)
    
    return {
        "id": new_comentario.id,
        "orden_trabajo_id": new_comentario.orden_trabajo_id,
        "comentario": new_comentario.comentario,
        "autor": new_comentario.autor,
        "fecha_creacion": new_comentario.fecha_creacion.isoformat() + "Z"
    }



# --- ENDPOINTS KPI & DASHBOARD ---

@app.get("/api/dashboard/stats")
def get_dashboard_stats(db: Session = Depends(get_db)):
    ots = db.exec(select(OrdenTrabajo)).all()
    
    total = len(ots)
    
    # Calcular estados mapeados
    creadas = 0
    asignadas = 0
    programadas = 0
    realizadas = 0
    
    for ot in ots:
        mapped_estado = ot.estado
        if mapped_estado in ("Resuelta", "REALIZADA"):
            mapped_estado = "REALIZADA"
        elif mapped_estado != "Cancelada":
            if ot.tecnico_id:
                if ot.fecha_programada:
                    mapped_estado = "PROGRAMADA"
                else:
                    mapped_estado = "ASIGNADA"
            else:
                mapped_estado = "CREADA"
                
        if mapped_estado == "CREADA":
            creadas += 1
        elif mapped_estado == "ASIGNADA":
            asignadas += 1
        elif mapped_estado == "PROGRAMADA":
            programadas += 1
        elif mapped_estado == "REALIZADA":
            realizadas += 1
            
    # Total activos y su estado
    activos = db.exec(select(Activo)).all()
    total_activos = len(activos)
    activos_operativos = len([a for a in activos if a.estado == "Operativo"])
    activos_falla = len([a for a in activos if a.estado == "En Reparación" or a.estado == "Inactivo"])
    
    # Cumplimiento preventivos vs correctivos
    preventivos = len([ot for ot in ots if ot.tipo == "Preventiva"])
    
    return {
        "kpis": {
            "total_ots": total,
            "creadas": creadas,
            "asignadas": asignadas,
            "programadas": programadas,
            "realizadas": realizadas,
            "total_activos": total_activos,
            "activos_operativos": activos_operativos,
            "activos_falla": activos_falla,
            "porcentaje_preventivo": round((preventivos / total * 100), 1) if total > 0 else 0.0
        }
    }

# --- ENDPOINTS PLANTILLAS DE CHEQUEO ---

@app.get("/api/plantillas", response_model=List[PlantillaChequeo])
def get_plantillas(db: Session = Depends(get_db)):
    return db.exec(select(PlantillaChequeo)).all()

@app.get("/api/plantillas/{plantilla_id}/items", response_model=List[ItemPlantillaChequeo])
def get_plantilla_items(plantilla_id: int, db: Session = Depends(get_db)):
    return db.exec(select(ItemPlantillaChequeo).where(ItemPlantillaChequeo.plantilla_id == plantilla_id)).all()

@app.post("/api/plantillas")
def create_plantilla(payload: dict, db: Session = Depends(get_db)):
    try:
        nombre = payload.get("nombre")
        descripcion = payload.get("descripcion")
        items_data = payload.get("items", [])
        
        if not nombre:
            raise HTTPException(status_code=400, detail="El nombre de la plantilla es obligatorio")
            
        plantilla = PlantillaChequeo(nombre=nombre, descripcion=descripcion)
        db.add(plantilla)
        db.commit()
        db.refresh(plantilla)
        
        for item in items_data:
            q_text = item.get("texto_pregunta")
            resp_type = item.get("tipo_respuesta", "booleano")
            unit = item.get("unidad_medida")
            
            if q_text:
                new_item = ItemPlantillaChequeo(
                    texto_pregunta=q_text,
                    tipo_respuesta=resp_type,
                    unidad_medida=unit,
                    plantilla_id=plantilla.id
                )
                db.add(new_item)
                
        db.commit()
        return {"status": "success", "id": plantilla.id, "nombre": plantilla.nombre}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/ordenes/{ot_id}/respuestas")
def get_orden_respuestas(ot_id: int, db: Session = Depends(get_db)):
    statement = select(RespuestaChequeo).where(RespuestaChequeo.orden_trabajo_id == ot_id)
    respuestas = db.exec(statement).all()
    
    results = []
    for r in respuestas:
        item = db.get(ItemPlantillaChequeo, r.item_plantilla_id)
        results.append({
            "id": r.id,
            "item_plantilla_id": r.item_plantilla_id,
            "texto_pregunta": item.texto_pregunta if item else "Pregunta eliminada",
            "tipo_respuesta": item.tipo_respuesta if item else "booleano",
            "unidad_medida": item.unidad_medida if item else None,
            "valor_booleano": r.valor_booleano,
            "valor_texto": r.valor_texto,
            "valor_numerico": r.valor_numerico,
            "observacion": r.observacion
        })
    return results

# --- ENDPOINTS EXCEL IMPORT / EXPORT ---

@app.get("/api/excel/exportar/ordenes")
def exportar_ordenes(db: Session = Depends(get_db)):
    query = select(OrdenTrabajo).order_by(OrdenTrabajo.id.desc())
    ots = db.exec(query).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ordenes de Trabajo"
    
    # Configurar zoom al 67%
    ws.sheet_view.zoomScale = 67
    
    headers = [
        "ID OT", "Planta", "Edificio", "Activo Afectado", "Técnico Asignado", "Reportado Por", "Ubicación",
        "Descripción", "Tipo Mantenimiento", "Estado", "Estado Ejecución", "Prioridad",
        "Fecha Creación", "Fecha Programada", "Fecha Inicio", "Fecha Realización", "Comentarios Técnicos"
    ]
    ws.append(headers)
    
    for ot in ots:
        planta = db.get(Planta, ot.planta_id)
        edificio = db.get(Edificio, ot.edificio_id)
        ubicacion = db.get(Ubicacion, ot.ubicacion_id) if ot.ubicacion_id else None
        activo = db.get(Activo, ot.activo_id) if ot.activo_id else None
        tecnico = db.get(Tecnico, ot.tecnico_id) if ot.tecnico_id else None
        
        # Mapeo dinámico del estado
        mapped_estado = ot.estado
        if mapped_estado in ("Resuelta", "REALIZADA"):
            mapped_estado = "REALIZADA"
        elif mapped_estado != "Cancelada":
            if ot.tecnico_id:
                if ot.fecha_programada:
                    mapped_estado = "PROGRAMADA"
                else:
                    mapped_estado = "ASIGNADA"
            else:
                mapped_estado = "CREADA"
                
        ws.append([
            f"OT-{ot.id}",
            planta.nombre if planta else "",
            edificio.nombre if edificio else "",
            activo.nombre if activo else "Reporte de Área",
            tecnico.nombre if tecnico else "No asignado",
            ot.reportado_por or "",
            ubicacion.nombre if ubicacion else "",
            ot.descripcion,
            ot.tipo,
            mapped_estado,
            ot.estado_ejecucion,
            ot.prioridad,
            ot.fecha_creacion.strftime("%y-%m-%d %H:%M") if ot.fecha_creacion else "",
            ot.fecha_programada.strftime("%y-%m-%d %H:%M") if ot.fecha_programada else "",
            ot.fecha_inicio.strftime("%y-%m-%d %H:%M") if ot.fecha_inicio else "",
            ot.fecha_resolucion.strftime("%y-%m-%d %H:%M") if ot.fecha_resolucion else "",
            ot.comentarios_tecnicos or ""
        ])
        
    from openpyxl.styles import Font, PatternFill
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        
    # Congelar primera fila
    ws.freeze_panes = 'A2'
    
    # Filtros de cabecera
    if ws.max_column > 0 and ws.max_row > 0:
        last_col_letter = openpyxl.utils.get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"
        
    # Aplicar anchos de columnas fijos del archivo de referencia (1020)
    column_widths = {
        "A": 8.63,   # ID OT
        "B": 10.91,  # Planta
        "C": 17.45,  # Edificio
        "D": 22.73,  # Activo Afectado
        "E": 4.0,    # Técnico Asignado
        "F": 10.63,  # Reportado Por
        "G": 28.54,  # Ubicación
        "H": 38.91,  # Descripción
        "I": 8.82,   # Tipo Mantenimiento
        "J": 6.91,   # Estado
        "K": 12.45,  # Estado Ejecución
        "L": 9.09,   # Prioridad
        "M": 12.63,  # Fecha Creación
        "N": 10.27,  # Fecha Programada
        "O": 12.0,   # Fecha Inicio
        "P": 12.09,  # Fecha Realización
        "Q": 26.27   # Comentarios Técnicos
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
        
    # Aplicar alineación vertical superior a toda la planilla y ajustar texto en Descripción (Columna H)
    from openpyxl.styles import Alignment
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.column_letter == "H" and cell.row > 1:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="top")
                
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reporte_ordenes.xlsx"}
    )

@app.get("/api/excel/exportar/activos")
def exportar_activos(db: Session = Depends(get_db)):
    activos = db.exec(select(Activo)).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario de Activos"
    
    headers = [
        "ID Activo", "Nombre", "Tipo", "Marca", "Modelo", "N° Serie", "Estado",
        "Planta", "Edificio", "Ubicación"
    ]
    ws.append(headers)
    
    for a in activos:
        ubicacion = db.get(Ubicacion, a.ubicacion_id) if a.ubicacion_id else None
        edificio = db.get(Edificio, ubicacion.edificio_id) if ubicacion else None
        planta = db.get(Planta, edificio.planta_id) if edificio else None
        
        ws.append([
            f"ACT-{a.id}",
            a.nombre,
            a.tipo,
            a.marca or "",
            a.modelo or "",
            a.numero_serie or "",
            a.estado,
            planta.nombre if planta else "",
            edificio.nombre if edificio else "",
            ubicacion.nombre if ubicacion else ""
        ])
        
    from openpyxl.styles import Font, PatternFill
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        
    # Congelar primera fila
    ws.freeze_panes = 'A2'
    
    # Filtros de cabecera
    if ws.max_column > 0 and ws.max_row > 0:
        last_col_letter = openpyxl.utils.get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventario_activos.xlsx"}
    )

@app.get("/api/excel/plantilla-importacion")
def descargar_plantilla_importacion(db: Session = Depends(get_db)):
    wb = openpyxl.Workbook()
    
    # Hoja 1: Plantilla
    ws_plantilla = wb.active
    ws_plantilla.title = "Importar Activos"
    headers_plantilla = [
        "Planta", "Edificio", "Ubicación", "Nombre Activo", "Tipo", "Marca", "Modelo", "N° Serie"
    ]
    ws_plantilla.append(headers_plantilla)
    
    # Consultar activos existentes en la base de datos (excluyendo inactivos)
    activos = db.exec(
        select(Activo)
        .where(Activo.estado != "Reemplazado")
        .where(Activo.estado != "Eliminado sin Reemplazo")
    ).all()
    
    activos_data = []
    for a in activos:
        ubicacion = db.get(Ubicacion, a.ubicacion_id) if a.ubicacion_id else None
        edificio = db.get(Edificio, ubicacion.edificio_id) if ubicacion else None
        planta = db.get(Planta, edificio.planta_id) if edificio else None
        activos_data.append({
            "planta": planta.nombre if planta else "",
            "edificio": edificio.nombre if edificio else "",
            "ubicacion": ubicacion.nombre if ubicacion else "",
            "nombre": a.nombre,
            "tipo": a.tipo,
            "marca": a.marca or "",
            "modelo": a.modelo or "",
            "serie": a.numero_serie or ""
        })
        
    # Ordenar los activos por planta, edificio, ubicacion y nombre
    activos_data.sort(key=lambda x: (x["planta"], x["edificio"], x["ubicacion"], x["nombre"]))
    
    if len(activos_data) > 0:
        for item in activos_data:
            ws_plantilla.append([
                item["planta"],
                item["edificio"],
                item["ubicacion"],
                item["nombre"],
                item["tipo"],
                item["marca"],
                item["modelo"],
                item["serie"]
            ])
    else:
        # Fila de ejemplo si la base de datos está vacía
        ws_plantilla.append([
            "Santa Adela",
            "Edificio Corporativo",
            "Oficina Gerencia [EC-P2-ON-OF1]",
            "Fancoil Muro Oficina 1",
            "Climatizacion",
            "Carrier",
            "42GW",
            "FC-SAMPLE-001"
        ])
    
    # Estilo de cabeceras de la Plantilla
    from openpyxl.styles import Font, PatternFill
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for col in range(1, len(headers_plantilla) + 1):
        cell = ws_plantilla.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
    
    # Hoja 2: Ubicaciones de Referencia
    ws_ubicaciones = wb.create_sheet(title="Ubicaciones de Referencia")
    headers_ubicaciones = ["Planta", "Edificio", "Ubicación (Nombre Exacto)"]
    ws_ubicaciones.append(headers_ubicaciones)
    
    # Obtener todas las ubicaciones ordenadas
    ubicaciones = db.exec(select(Ubicacion)).all()
    locs_data = []
    for u in ubicaciones:
        ed = db.get(Edificio, u.edificio_id) if u.edificio_id else None
        pl = db.get(Planta, ed.planta_id) if ed else None
        locs_data.append({
            "planta": pl.nombre if pl else "",
            "edificio": ed.nombre if ed else "",
            "ubicacion": u.nombre
        })
        
    # Ordenar por planta, edificio y ubicacion
    locs_data.sort(key=lambda x: (x["planta"], x["edificio"], x["ubicacion"]))
    
    for item in locs_data:
        ws_ubicaciones.append([item["planta"], item["edificio"], item["ubicacion"]])
        
    # Estilo de cabeceras de Ubicaciones
    for col in range(1, len(headers_ubicaciones) + 1):
        cell = ws_ubicaciones.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        
    # Ajuste automático del ancho de columna, filtros y congelamiento
    for ws in [ws_plantilla, ws_ubicaciones]:
        # Congelar primera fila
        ws.freeze_panes = 'A2'
        
        # Filtros de cabecera
        if ws.max_column > 0 and ws.max_row > 0:
            last_col_letter = openpyxl.utils.get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_importar_activos.xlsx"}
    )

@app.post("/api/excel/importar/activos")

async def importar_activos(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Por favor sube un archivo Excel válido (.xlsx)")
        
    try:
        contents = await file.read()
        file_stream = io.BytesIO(contents)
        wb = openpyxl.load_workbook(file_stream, read_only=True)
        if "Importar Activos" in wb.sheetnames:
            sheet = wb["Importar Activos"]
        elif "Plantilla Importar Activos" in wb.sheetnames:
            sheet = wb["Plantilla Importar Activos"]
        else:
            sheet = wb.active
        
        imported_count = 0
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            return {"status": "success", "imported": 0, "message": "El archivo está vacío o no contiene filas de datos"}
            
        seen_in_excel = {}
        for row in rows[1:]:
            if not row or not any(row):
                continue
                
            planta_name = str(row[0]).strip() if row[0] is not None else ""
            edificio_name = str(row[1]).strip() if row[1] is not None else ""
            ubicacion_name = str(row[2]).strip() if row[2] is not None else ""
            activo_name = str(row[3]).strip() if row[3] is not None else ""
            tipo_val = str(row[4]).strip() if row[4] is not None else "Climatizacion"
            marca_val = str(row[5]).strip() if row[5] is not None else None
            modelo_val = str(row[6]).strip() if row[6] is not None else None
            serie_val = str(row[7]).strip() if row[7] is not None else None
            
            if not planta_name or not edificio_name or not ubicacion_name or not activo_name:
                continue
                
            planta = db.exec(select(Planta).where(Planta.nombre == planta_name)).first()
            if not planta:
                planta = Planta(nombre=planta_name)
                db.add(planta)
                db.commit()
                db.refresh(planta)
                
            edificio = db.exec(select(Edificio).where(Edificio.nombre == edificio_name, Edificio.planta_id == planta.id)).first()
            if not edificio:
                edificio = Edificio(nombre=edificio_name, planta_id=planta.id)
                db.add(edificio)
                db.commit()
                db.refresh(edificio)
                
            ubicacion = db.exec(select(Ubicacion).where(Ubicacion.nombre == ubicacion_name, Ubicacion.edificio_id == edificio.id)).first()
            if not ubicacion:
                statement = select(Ubicacion).where(Ubicacion.edificio_id == edificio.id)
                all_ubics = db.exec(statement).all()
                matched_u = None
                for u in all_ubics:
                    if u.nombre.startswith(ubicacion_name):
                        matched_u = u
                        break
                if matched_u:
                    ubicacion = matched_u
                else:
                    ubicacion = Ubicacion(nombre=ubicacion_name, edificio_id=edificio.id)
                    db.add(ubicacion)
                    db.commit()
                    db.refresh(ubicacion)
                    
            key = (activo_name, serie_val, ubicacion.id)
            seen_in_excel[key] = seen_in_excel.get(key, 0) + 1
            
            db_count = 0
            if serie_val:
                db_count = len(db.exec(
                    select(Activo)
                    .where(Activo.numero_serie == serie_val)
                    .where(Activo.ubicacion_id == ubicacion.id)
                    .where(Activo.estado != "Reemplazado")
                    .where(Activo.estado != "Eliminado sin Reemplazo")
                ).all())
            else:
                db_count = len(db.exec(
                    select(Activo)
                    .where(Activo.nombre == activo_name)
                    .where(Activo.ubicacion_id == ubicacion.id)
                    .where(Activo.estado != "Reemplazado")
                    .where(Activo.estado != "Eliminado sin Reemplazo")
                ).all())
                
            if seen_in_excel[key] > db_count:
                activo = Activo(
                    nombre=activo_name,
                    tipo=tipo_val,
                    marca=marca_val,
                    modelo=modelo_val,
                    numero_serie=serie_val,
                    estado="Operativo",
                    ubicacion_id=ubicacion.id
                )
                db.add(activo)
                db.commit()
                imported_count += 1
                
        return {"status": "success", "imported": imported_count, "message": f"Se importaron {imported_count} activos correctamente."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al procesar archivo Excel: {str(e)}")

@app.get("/api/excel/plantilla-despiece")
def descargar_plantilla_despiece(db: Session = Depends(get_db)):
    wb = openpyxl.Workbook()
    
    # Hoja 1: Plantilla de importación
    ws_import = wb.active
    ws_import.title = "Importar Despieces"
    
    headers_import = [
        "ID Componente (Solo para actualizar, ej: PZ-123)",
        "N° Serie Activo * (Recomendado)", 
        "Planta (Opcional)", 
        "Edificio (Opcional)", 
        "Ubicación (Opcional)", 
        "Nombre Activo (Opcional)", 
        "Nombre Componente * (Obligatorio)", 
        "Marca", 
        "Modelo", 
        "N° Serie Componente / Obs", 
        "Estado"
    ]
    ws_import.append(headers_import)
    
    # Consultar componentes y sus activos para pre-rellenar plantilla
    comps = db.exec(
        select(ComponenteActivo)
        .join(Activo)
        .where(Activo.estado != "Reemplazado")
        .where(Activo.estado != "Eliminado sin Reemplazo")
    ).all()
    
    comps_data = []
    for c in comps:
        activo = db.get(Activo, c.activo_id)
        if not activo:
            continue
        ubicacion = db.get(Ubicacion, activo.ubicacion_id) if activo.ubicacion_id else None
        edificio = db.get(Edificio, ubicacion.edificio_id) if ubicacion else None
        planta = db.get(Planta, edificio.planta_id) if edificio else None
        
        comps_data.append({
            "comp": c,
            "activo_serie": activo.numero_serie or "",
            "planta": planta.nombre if planta else "",
            "edificio": edificio.nombre if edificio else "",
            "ubicacion": ubicacion.nombre if ubicacion else "",
            "activo_nombre": activo.nombre
        })
        
    # Ordenar por planta, edificio, ubicacion, activo, nombre componente
    comps_data.sort(key=lambda x: (x["planta"], x["edificio"], x["ubicacion"], x["activo_nombre"], x["comp"].nombre))
    
    if len(comps_data) > 0:
        for item in comps_data:
            c = item["comp"]
            ws_import.append([
                f"PZ-{c.id}",
                item["activo_serie"],
                item["planta"],
                item["edificio"],
                item["ubicacion"],
                item["activo_nombre"],
                c.nombre,
                c.marca or "",
                c.modelo or "",
                c.numero_serie or "",
                c.estado
            ])
    else:
        # Fila de ejemplo si no hay componentes en la DB
        ws_import.append([
            "PZ-1",
            "FC-P1A-E1",
            "Santa Adela",
            "Edificio Corporativo",
            "Oficina RRHH [EC-P1-RecN-#1]",
            "Fancoil FC-P1A-E1",
            "Compresor",
            "Carrier",
            "42GW",
            "Obs: Buen estado",
            "Operativo"
        ])
    
    # Estilos cabecera Hoja 1
    from openpyxl.styles import Font, PatternFill
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for col in range(1, len(headers_import) + 1):
        cell = ws_import.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        
    # Hoja 2: Activos de Referencia
    ws_ref = wb.create_sheet(title="Activos de Referencia")
    headers_ref = [
        "ID Activo", "Nombre Activo", "N° Serie Activo", "Planta", "Edificio", "Ubicación"
    ]
    ws_ref.append(headers_ref)
    
    activos = db.exec(
        select(Activo)
        .where(Activo.estado != "Reemplazado")
        .where(Activo.estado != "Eliminado sin Reemplazo")
    ).all()
    
    # Resolver ubicaciones de referencia
    activos_ref_data = []
    for a in activos:
        ubicacion = db.get(Ubicacion, a.ubicacion_id) if a.ubicacion_id else None
        edificio = db.get(Edificio, ubicacion.edificio_id) if ubicacion else None
        planta = db.get(Planta, edificio.planta_id) if edificio else None
        
        activos_ref_data.append([
            f"ACT-{a.id}",
            a.nombre,
            a.numero_serie or "",
            planta.nombre if planta else "",
            edificio.nombre if edificio else "",
            ubicacion.nombre if ubicacion else ""
        ])
        
    activos_ref_data.sort(key=lambda x: (x[3], x[4], x[5], x[1]))
    for row in activos_ref_data:
        ws_ref.append(row)
        
    # Estilos cabecera Hoja 2
    for col in range(1, len(headers_ref) + 1):
        cell = ws_ref.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        
    # Autoajuste de columnas, filtros y congelamiento
    for ws in [ws_import, ws_ref]:
        # Congelar primera fila
        ws.freeze_panes = 'A2'
        
        # Filtros de cabecera
        if ws.max_column > 0 and ws.max_row > 0:
            last_col_letter = openpyxl.utils.get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_importar_despiece.xlsx"}
    )

@app.post("/api/excel/importar/despiece")
async def importar_despiece(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Por favor sube un archivo Excel válido (.xlsx)")
        
    try:
        contents = await file.read()
        file_stream = io.BytesIO(contents)
        wb = openpyxl.load_workbook(file_stream, read_only=True)
        
        # Seleccionar hoja
        if "Importar Despieces" in wb.sheetnames:
            sheet = wb["Importar Despieces"]
        else:
            sheet = wb.active
            
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            return {"status": "success", "imported": 0, "message": "El archivo está vacío o no contiene filas de datos"}
            
        imported_count = 0
        updated_count = 0
        omitted_count = 0
        unmatched_rows = []
        
        for idx, row in enumerate(rows[1:], start=2):
            if not row or not any(row):
                continue
                
            # Leer columnas
            id_comp_str = str(row[0]).strip() if row[0] is not None else ""
            serie_activo = str(row[1]).strip() if row[1] is not None else ""
            planta_name = str(row[2]).strip() if row[2] is not None else ""
            edificio_name = str(row[3]).strip() if row[3] is not None else ""
            ubicacion_name = str(row[4]).strip() if row[4] is not None else ""
            activo_name = str(row[5]).strip() if row[5] is not None else ""
            comp_name = str(row[6]).strip() if row[6] is not None else ""
            marca_val = str(row[7]).strip() if row[7] is not None else None
            modelo_val = str(row[8]).strip() if row[8] is not None else None
            comp_serie = str(row[9]).strip() if row[9] is not None else None
            estado_val = str(row[10]).strip() if row[10] is not None else "Operativo"
            
            # Fila de ejemplo (ejemplo Compresor Carrier 42GW) se omite
            if comp_name == "Compresor" and (serie_activo == "FC-P1A-E1" or id_comp_str == "PZ-1"):
                continue
                
            if not comp_name:
                continue
                
            # Limpiar ID si se provee
            comp_id = None
            if id_comp_str:
                clean_id = id_comp_str.upper().replace("PZ-", "").replace("PZ", "").strip()
                if clean_id.isdigit():
                    comp_id = int(clean_id)
                    
            # Si se provee ID, intentar actualizar
            existing_comp = None
            if comp_id:
                existing_comp = db.get(ComponenteActivo, comp_id)
                
            if existing_comp:
                # Buscar y asociar activo
                activo = None
                if serie_activo:
                    activo = db.exec(
                        select(Activo)
                        .where(Activo.numero_serie == serie_activo)
                        .where(Activo.estado != "Reemplazado")
                        .where(Activo.estado != "Eliminado sin Reemplazo")
                    ).first()
                    
                if not activo and activo_name and planta_name and edificio_name and ubicacion_name:
                    planta = db.exec(select(Planta).where(Planta.nombre == planta_name)).first()
                    if planta:
                        edificio = db.exec(select(Edificio).where(Edificio.nombre == edificio_name, Edificio.planta_id == planta.id)).first()
                        if edificio:
                            ubicacion = db.exec(select(Ubicacion).where(Ubicacion.nombre == ubicacion_name, Ubicacion.edificio_id == edificio.id)).first()
                            if ubicacion:
                                activo = db.exec(
                                    select(Activo)
                                    .where(Activo.nombre == activo_name)
                                    .where(Activo.ubicacion_id == ubicacion.id)
                                    .where(Activo.estado != "Reemplazado")
                                    .where(Activo.estado != "Eliminado sin Reemplazo")
                                ).first()
                
                # Actualizar datos
                existing_comp.nombre = comp_name
                existing_comp.marca = marca_val
                existing_comp.modelo = modelo_val
                existing_comp.numero_serie = comp_serie
                existing_comp.estado = estado_val
                if activo:
                    existing_comp.activo_id = activo.id
                    
                db.add(existing_comp)
                updated_count += 1
                continue
                
            # Si se especificó un ID pero no se encontró en base de datos, lo tratamos como nuevo
            if comp_id and not existing_comp:
                pass
                
            # Buscar activo para nuevo componente
            activo = None
            if serie_activo:
                activo = db.exec(
                    select(Activo)
                    .where(Activo.numero_serie == serie_activo)
                    .where(Activo.estado != "Reemplazado")
                    .where(Activo.estado != "Eliminado sin Reemplazo")
                ).first()
                
            if not activo and activo_name and planta_name and edificio_name and ubicacion_name:
                planta = db.exec(select(Planta).where(Planta.nombre == planta_name)).first()
                if planta:
                    edificio = db.exec(select(Edificio).where(Edificio.nombre == edificio_name, Edificio.planta_id == planta.id)).first()
                    if edificio:
                        ubicacion = db.exec(select(Ubicacion).where(Ubicacion.nombre == ubicacion_name, Ubicacion.edificio_id == edificio.id)).first()
                        if ubicacion:
                            activo = db.exec(
                                select(Activo)
                                .where(Activo.nombre == activo_name)
                                .where(Activo.ubicacion_id == ubicacion.id)
                                .where(Activo.estado != "Reemplazado")
                                .where(Activo.estado != "Eliminado sin Reemplazo")
                            ).first()
                            
            if not activo:
                unmatched_rows.append(f"Fila {idx}: no se encontró activo para Serie '{serie_activo}' o Nombre '{activo_name}'")
                continue
                
            # Verificar duplicado de componente en el mismo activo
            existing_comps = db.exec(
                select(ComponenteActivo)
                .where(ComponenteActivo.activo_id == activo.id)
                .where(ComponenteActivo.nombre == comp_name)
            ).all()
            
            is_duplicate = False
            for ec in existing_comps:
                if ec.modelo == modelo_val and ec.numero_serie == comp_serie:
                    is_duplicate = True
                    break
                    
            if is_duplicate:
                omitted_count += 1
                continue
                
            # Insertar componente
            new_comp = ComponenteActivo(
                nombre=comp_name,
                marca=marca_val,
                modelo=modelo_val,
                numero_serie=comp_serie,
                estado=estado_val,
                activo_id=activo.id
            )
            db.add(new_comp)
            imported_count += 1
            
        db.commit()
        
        msg = f"Se importaron {imported_count} componentes y se actualizaron {updated_count} correctamente. Se omitieron {omitted_count} duplicados."
        if unmatched_rows:
            msg += f" Omitidos {len(unmatched_rows)} por no encontrados: {', '.join(unmatched_rows[:3])}..."
            
        return {
            "status": "success",
            "imported": imported_count,
            "updated": updated_count,
            "omitted": omitted_count,
            "message": msg
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al procesar archivo Excel: {str(e)}")

@app.get("/api/excel/plantilla-checklists")
def descargar_plantilla_checklists(db: Session = Depends(get_db)):
    wb = openpyxl.Workbook()
    
    # Hoja 1: Importar Checklists
    ws_import = wb.active
    ws_import.title = "Importar Checklists"
    
    headers = [
        "ID Plantilla (Solo para actualizar, ej: PL-12)",
        "Nombre de la Plantilla * (Obligatorio)",
        "Descripción de la Plantilla",
        "Tipo de Revisión * (Ronda o Chequeo Preventivo)",
        "Pregunta / Ítem * (Obligatorio)",
        "Tipo de Respuesta * (booleano, numerico, o texto)",
        "Unidad de Medida (Opcional, ej: V, A, psi, C)",
        "Activo Asociado (Opcional, ej: 304 - Fancoil 5 o dejar vacío)"
    ]
    ws_import.append(headers)
    
    # Consultar plantillas existentes para pre-llenar plantilla
    plantillas = db.exec(select(PlantillaChequeo)).all()
    
    rows_written = 0
    for p in plantillas:
        items = db.exec(select(ItemPlantillaChequeo).where(ItemPlantillaChequeo.plantilla_id == p.id)).all()
        
        activo_info = ""
        if p.activo_id:
            activo = db.get(Activo, p.activo_id)
            if activo:
                activo_info = f"{activo.id} - {activo.nombre}"
            else:
                activo_info = str(p.activo_id)

        for item in items:
            ws_import.append([
                f"PL-{p.id}",
                p.nombre,
                p.descripcion or "",
                p.tipo_revision or "Chequeo Preventivo",
                item.texto_pregunta,
                item.tipo_respuesta,
                item.unidad_medida or "",
                activo_info
            ])
            rows_written += 1
            
    if rows_written == 0:
        # Fila de ejemplo
        ws_import.append([
            "PL-Ejemplo",
            "Mantenimiento Preventivo Fancoil",
            "Inspección de mantención mensual de fancoils",
            "Chequeo Preventivo",
            "Limpieza de filtros de aire de unidad evaporadora",
            "booleano",
            "",
            "304 - Fancoil 5"
        ])
        ws_import.append([
            "PL-Ejemplo",
            "Mantenimiento Preventivo Fancoil",
            "Inspección de mantención mensual de fancoils",
            "Chequeo Preventivo",
            "Medición de consumo eléctrico del motor",
            "numerico",
            "A",
            "304 - Fancoil 5"
        ])
        ws_import.append([
            "PL-Ejemplo",
            "Ronda Diaria de Calderas",
            "Recorrido visual diario por sala de calderas",
            "Ronda",
            "Comprobar fugas de agua o vapor",
            "booleano",
            "",
            ""
        ])
        
    # Hoja 2: Instrucciones
    ws_instr = wb.create_sheet(title="Instrucciones")
    ws_instr.append(["Guía de Llenado para Importación de Checklists"])
    ws_instr.append([])
    ws_instr.append(["1. Columnas obligatorias:"])
    ws_instr.append(["   - Nombre de la Plantilla: El nombre que identifica al checklist entero."])
    ws_instr.append(["   - Tipo de Revisión: Debe ser exacto: 'Ronda' o 'Chequeo Preventivo'."])
    ws_instr.append(["   - Pregunta / Ítem: La pregunta de la inspección."])
    ws_instr.append(["   - Tipo de Respuesta: Debe ser: 'booleano' (si es OK/Falla), 'numerico' (para ingresar un valor decimal) o 'texto' (para notas)."])
    ws_instr.append([])
    ws_instr.append(["2. Modificación de Plantillas Existentes:"])
    ws_instr.append(["   - Para modificar o añadir ítems a una plantilla existente, mantén el 'ID Plantilla' (ej. PL-1)."])
    ws_instr.append(["   - Si dejas el 'ID Plantilla' vacío, el sistema buscará la plantilla por su nombre, y si no existe la creará."])
    
    # Autoajuste de columnas, filtros y congelamiento
    for ws in [ws_import]:
        ws.freeze_panes = 'A2'
        if ws.max_column > 0 and ws.max_row > 0:
            last_col_letter = openpyxl.utils.get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"
            
    for ws in [ws_import, ws_instr]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Stream file
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    headers_response = {
        'Content-Disposition': 'attachment; filename="plantilla_importar_checklists.xlsx"'
    }
    return StreamingResponse(
        buffer,
        headers=headers_response,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.post("/api/excel/importar/checklists")
async def importar_checklists(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Por favor sube un archivo Excel válido (.xlsx)")
        
    try:
        contents = await file.read()
        file_stream = io.BytesIO(contents)
        wb = openpyxl.load_workbook(file_stream, read_only=True)
        
        # Seleccionar hoja
        if "Importar Checklists" in wb.sheetnames:
            sheet = wb["Importar Checklists"]
        else:
            sheet = wb.active
            
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            return {"status": "success", "imported_templates": 0, "imported_items": 0, "message": "El archivo está vacío o no contiene filas de datos"}
            
        templates_created = 0
        templates_updated = 0
        items_created = 0
        items_updated = 0
        
        resolved_templates = {}
        processed_templates = set()
        
        for idx, row in enumerate(rows[1:], start=2):
            if not row or not any(row):
                continue
                
            id_plantilla_str = str(row[0]).strip() if row[0] is not None else ""
            nombre_plantilla = str(row[1]).strip() if row[1] is not None else ""
            descripcion_plantilla = str(row[2]).strip() if row[2] is not None else ""
            tipo_rev_val = str(row[3]).strip() if row[3] is not None else "Chequeo Preventivo"
            pregunta = str(row[4]).strip() if row[4] is not None else ""
            tipo_resp_val = str(row[5]).strip() if row[5] is not None else "booleano"
            unidad_medida = str(row[6]).strip() if row[6] is not None else ""
            activo_asociado_str = str(row[7]).strip() if len(row) > 7 and row[7] is not None else ""
            
            # Omitir filas de ejemplo
            if id_plantilla_str == "PL-Ejemplo" or nombre_plantilla == "Mantenimiento Preventivo Fancoil" and pregunta == "Limpieza de filtros de aire de unidad evaporadora":
                continue
                
            if not nombre_plantilla or not pregunta:
                continue
                
            # Normalizar tipo de revision
            if tipo_rev_val.lower() in ["ronda", "rondas"]:
                tipo_rev = "Ronda"
            else:
                tipo_rev = "Chequeo Preventivo"
                
            # Normalizar tipo de respuesta
            if tipo_resp_val.lower() in ["numerico", "numérico", "numero", "número"]:
                tipo_resp = "numerico"
            elif tipo_resp_val.lower() in ["texto", "string", "observaciones"]:
                tipo_resp = "texto"
            else:
                tipo_resp = "booleano"
                
            plantilla = None
            
            # 1. Por ID si se provee
            if id_plantilla_str and id_plantilla_str.upper() not in ["PL-EJEMPLO", ""]:
                clean_id = id_plantilla_str.upper().replace("PL-", "").replace("PL", "").strip()
                if clean_id.isdigit():
                    p_id = int(clean_id)
                    plantilla = db.get(PlantillaChequeo, p_id)
                    
            # 2. Si no se encontró o no se proveyó ID, buscar por nombre
            if not plantilla:
                if nombre_plantilla in resolved_templates:
                    plantilla = resolved_templates[nombre_plantilla]
                else:
                    plantilla = db.exec(select(PlantillaChequeo).where(PlantillaChequeo.nombre == nombre_plantilla)).first()
            
            # Resolve Asset
            activo_id = None
            if activo_asociado_str:
                parts = [p.strip() for p in activo_asociado_str.split("-", 1)]
                if parts[0].isdigit():
                    act_id = int(parts[0])
                    activo = db.get(Activo, act_id)
                    if activo:
                        activo_id = activo.id
                
                if not activo_id:
                    activo = db.exec(select(Activo).where(Activo.nombre == activo_asociado_str)).first()
                    if activo:
                        activo_id = activo.id
                    elif len(parts) > 1:
                        activo = db.exec(select(Activo).where(Activo.nombre == parts[1])).first()
                        if activo:
                            activo_id = activo.id

            if plantilla:
                resolved_templates[nombre_plantilla] = plantilla
                if plantilla.id not in processed_templates:
                    has_changed = False
                    if descripcion_plantilla and plantilla.descripcion != descripcion_plantilla:
                        plantilla.descripcion = descripcion_plantilla
                        has_changed = True
                    if tipo_rev and plantilla.tipo_revision != tipo_rev:
                        plantilla.tipo_revision = tipo_rev
                        has_changed = True
                    
                    # Update asset association
                    if activo_asociado_str:
                        if plantilla.activo_id != activo_id:
                            plantilla.activo_id = activo_id
                            has_changed = True
                    else:
                        if plantilla.activo_id is not None:
                            plantilla.activo_id = None
                            has_changed = True
                            
                    if has_changed:
                        db.add(plantilla)
                        db.commit()
                        db.refresh(plantilla)
                        templates_updated += 1
                    processed_templates.add(plantilla.id)
            else:
                plantilla = PlantillaChequeo(
                    nombre=nombre_plantilla,
                    descripcion=descripcion_plantilla or None,
                    tipo_revision=tipo_rev,
                    activo_id=activo_id
                )
                db.add(plantilla)
                db.commit()
                db.refresh(plantilla)
                templates_created += 1
                resolved_templates[nombre_plantilla] = plantilla
                processed_templates.add(plantilla.id)
                
            # Procesar el item de la plantilla
            existing_item = db.exec(
                select(ItemPlantillaChequeo)
                .where(ItemPlantillaChequeo.plantilla_id == plantilla.id)
                .where(ItemPlantillaChequeo.texto_pregunta == pregunta)
            ).first()
            
            if existing_item:
                has_item_changed = False
                if existing_item.tipo_respuesta != tipo_resp:
                    existing_item.tipo_respuesta = tipo_resp
                    has_item_changed = True
                if unidad_medida and existing_item.unidad_medida != unidad_medida:
                    existing_item.unidad_medida = unidad_medida
                    has_item_changed = True
                elif not unidad_medida and existing_item.unidad_medida:
                    existing_item.unidad_medida = None
                    has_item_changed = True
                    
                if has_item_changed:
                    db.add(existing_item)
                    db.commit()
                    items_updated += 1
            else:
                new_item = ItemPlantillaChequeo(
                    texto_pregunta=pregunta,
                    tipo_respuesta=tipo_resp,
                    unidad_medida=unidad_medida if unidad_medida else None,
                    plantilla_id=plantilla.id
                )
                db.add(new_item)
                db.commit()
                items_created += 1
                
        db.commit()
        
        msg = f"Importación exitosa. Plantillas: {templates_created} creadas, {templates_updated} actualizadas. Preguntas: {items_created} creadas, {items_updated} actualizadas."
        return {
            "status": "success",
            "imported_templates": templates_created + templates_updated,
            "imported_items": items_created + items_updated,
            "message": msg
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al procesar el archivo Excel: {str(e)}")

def serialize_cell_color(cell) -> Optional[str]:
    fill = cell.fill
    if not fill or not fill.fill_type or fill.fill_type == 'none':
        return None
    sc = fill.start_color
    if not sc or sc.type is None:
        return None
    if sc.type == 'rgb' and sc.rgb and sc.rgb != "00000000" and sc.rgb != "FFFFFFFF":
        return f"rgb:{sc.rgb}"
    elif sc.type == 'theme' and sc.theme is not None:
        tint = sc.tint if sc.tint is not None else 0.0
        return f"theme:{sc.theme}:{tint}"
    elif sc.type == 'indexed' and sc.indexed is not None:
        return f"indexed:{sc.indexed}"
    return None

def deserialize_and_apply_color(cell, color_str: Optional[str]):
    if not color_str:
        return
    from openpyxl.styles import PatternFill
    from openpyxl.styles.colors import Color
    
    parts = color_str.split(":")
    if not parts or len(parts) < 2:
        return
        
    color_type = parts[0]
    try:
        if color_type == "rgb":
            rgb_val = parts[1]
            cell.fill = PatternFill(start_color=rgb_val, end_color=rgb_val, fill_type="solid")
        elif color_type == "theme" and len(parts) >= 3:
            theme_val = int(parts[1])
            tint_val = float(parts[2])
            c = Color(type="theme", theme=theme_val, tint=tint_val)
            cell.fill = PatternFill(start_color=c, end_color=c, fill_type="solid")
        elif color_type == "indexed":
            indexed_val = int(parts[1])
            c = Color(type="indexed", indexed=indexed_val)
            cell.fill = PatternFill(start_color=c, end_color=c, fill_type="solid")
    except Exception as e:
        print(f"Error applying color {color_str} to cell: {e}")

@app.get("/api/excel/plantilla-unificada")
def descargar_plantilla_unificada(db: Session = Depends(get_db)):
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # HOJA 1: Importar Ubicaciones (Editable)
    # -------------------------------------------------------------
    ws_ubicaciones = wb.active
    ws_ubicaciones.title = "Importar Ubicaciones"
    
    headers_ubicaciones = [
        "Planta *", "Edificio *", "Ubicación (Nombre Completo) *", "Código", "Uso", "Cargo", "Ocupantes"
    ]
    ws_ubicaciones.append(headers_ubicaciones)
    
    ubicaciones = db.exec(select(Ubicacion)).all()
    ref_locs_data = []
    for u in ubicaciones:
        edificio = db.get(Edificio, u.edificio_id) if u.edificio_id else None
        planta = db.get(Planta, edificio.planta_id) if edificio else None
        ocupantes_str = ", ".join([o.nombre for o in u.ocupantes])
        ref_locs_data.append({
            "row": [
                planta.nombre if planta else "",
                edificio.nombre if edificio else "",
                u.nombre,
                u.codigo or "",
                u.uso or "Oficina",
                u.cargo or "",
                ocupantes_str
            ],
            "color": u.color
        })
    ref_locs_data.sort(key=lambda x: (x["row"][0], x["row"][1], x["row"][2]))
    for item in ref_locs_data:
        ws_ubicaciones.append(item["row"])
        if item["color"]:
            row_idx = ws_ubicaciones.max_row
            for col_idx in range(1, len(item["row"]) + 1):
                cell = ws_ubicaciones.cell(row=row_idx, column=col_idx)
                deserialize_and_apply_color(cell, item["color"])
        
    # Estilos cabecera Hoja 1
    from openpyxl.styles import Font, PatternFill
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill_ubic = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
    for col in range(1, len(headers_ubicaciones) + 1):
        cell = ws_ubicaciones.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill_ubic
        
    # -------------------------------------------------------------
    # HOJA 2: Importar Activos (Editable)
    # -------------------------------------------------------------
    ws_activos = wb.create_sheet(title="Importar Activos")
    headers_activos = [
        "Planta *", "Edificio *", "Ubicación *", "Nombre Activo *", "Tipo *", "Marca", "Modelo", "N° Serie"
    ]
    ws_activos.append(headers_activos)
    
    # Pre-llenar con activos operativos
    activos = db.exec(
        select(Activo)
        .where(Activo.estado != "Reemplazado")
        .where(Activo.estado != "Eliminado sin Reemplazo")
    ).all()
    
    activos_data = []
    for a in activos:
        ubicacion = db.get(Ubicacion, a.ubicacion_id) if a.ubicacion_id else None
        edificio = db.get(Edificio, ubicacion.edificio_id) if ubicacion else None
        planta = db.get(Planta, edificio.planta_id) if edificio else None
        
        activos_data.append({
            "planta": planta.nombre if planta else "",
            "edificio": edificio.nombre if edificio else "",
            "ubicacion": ubicacion.nombre if ubicacion else "",
            "nombre": a.nombre,
            "tipo": a.tipo,
            "marca": a.marca or "",
            "modelo": a.modelo or "",
            "serie": a.numero_serie or "",
            "color": a.color
        })
    activos_data.sort(key=lambda x: (x["planta"], x["edificio"], x["ubicacion"], x["nombre"]))
    
    for item in activos_data:
        ws_activos.append([
            item["planta"],
            item["edificio"],
            item["ubicacion"],
            item["nombre"],
            item["tipo"],
            item["marca"],
            item["modelo"],
            item["serie"]
        ])
        if item["color"]:
            row_idx = ws_activos.max_row
            for col_idx in range(1, len(headers_activos) + 1):
                cell = ws_activos.cell(row=row_idx, column=col_idx)
                deserialize_and_apply_color(cell, item["color"])
        
    # Estilos cabecera Hoja 2
    header_fill_act = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for col in range(1, len(headers_activos) + 1):
        cell = ws_activos.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill_act
        
    # -------------------------------------------------------------
    # HOJA 3: Importar Despieces (Editable)
    # -------------------------------------------------------------
    ws_despieces = wb.create_sheet(title="Importar Despieces")
    headers_despieces = [
        "ID Componente (Solo para actualizar, ej: PZ-123)",
        "N° Serie Activo * (Recomendado)", 
        "Planta (Opcional)", 
        "Edificio (Opcional)", 
        "Ubicación (Opcional)", 
        "Nombre Activo (Opcional)", 
        "Nombre Componente * (Obligatorio)", 
        "Marca", 
        "Modelo", 
        "N° Serie Componente / Obs", 
        "Estado"
    ]
    ws_despieces.append(headers_despieces)
    
    comps = db.exec(
        select(ComponenteActivo)
        .join(Activo)
        .where(Activo.estado != "Reemplazado")
        .where(Activo.estado != "Eliminado sin Reemplazo")
    ).all()
    
    comps_data = []
    for c in comps:
        activo = db.get(Activo, c.activo_id)
        if not activo:
            continue
        ubicacion = db.get(Ubicacion, activo.ubicacion_id) if activo.ubicacion_id else None
        edificio = db.get(Edificio, ubicacion.edificio_id) if ubicacion else None
        planta = db.get(Planta, edificio.planta_id) if edificio else None
        
        comps_data.append({
            "comp": c,
            "activo_serie": activo.numero_serie or "",
            "planta": planta.nombre if planta else "",
            "edificio": edificio.nombre if edificio else "",
            "ubicacion": ubicacion.nombre if ubicacion else "",
            "activo_nombre": activo.nombre
        })
    comps_data.sort(key=lambda x: (x["planta"], x["edificio"], x["ubicacion"], x["activo_nombre"], x["comp"].nombre))
    
    for item in comps_data:
        c = item["comp"]
        ws_despieces.append([
            f"PZ-{c.id}",
            item["activo_serie"],
            item["planta"],
            item["edificio"],
            item["ubicacion"],
            item["activo_nombre"],
            c.nombre,
            c.marca or "",
            c.modelo or "",
            c.numero_serie or "",
            c.estado
        ])
        
    # Estilos cabecera Hoja 3
    header_fill_desp = PatternFill(start_color="365F91", end_color="365F91", fill_type="solid")
    for col in range(1, len(headers_despieces) + 1):
        cell = ws_despieces.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill_desp
        
    # -------------------------------------------------------------
    # HOJA 4: Activos de Referencia (Solo Lectura)
    # -------------------------------------------------------------
    ws_ref_act = wb.create_sheet(title="Activos de Referencia")
    headers_ref_act = ["ID Activo", "Nombre Activo", "N° Serie Activo", "Planta", "Edificio", "Ubicación"]
    ws_ref_act.append(headers_ref_act)
    
    activos_ref_data = []
    for a in activos:
        ubicacion = db.get(Ubicacion, a.ubicacion_id) if a.ubicacion_id else None
        edificio = db.get(Edificio, ubicacion.edificio_id) if ubicacion else None
        planta = db.get(Planta, edificio.planta_id) if edificio else None
        
        activos_ref_data.append([
            f"ACT-{a.id}",
            a.nombre,
            a.numero_serie or "",
            planta.nombre if planta else "",
            edificio.nombre if edificio else "",
            ubicacion.nombre if ubicacion else ""
        ])
    activos_ref_data.sort(key=lambda x: (x[3], x[4], x[5], x[1]))
    for row in activos_ref_data:
        ws_ref_act.append(row)
        
    # Estilos cabecera Hoja 4
    header_fill_ref_act = PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid")
    for col in range(1, len(headers_ref_act) + 1):
        cell = ws_ref_act.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill_ref_act
        
    # Autoajuste de columnas, filtros y congelamiento de cabecera en todas las hojas
    for ws in [ws_ubicaciones, ws_activos, ws_despieces, ws_ref_act]:
        # Congelar primera fila
        ws.freeze_panes = 'A2'
        
        # Filtros de cabecera
        if ws.max_column > 0 and ws.max_row > 0:
            last_col_letter = openpyxl.utils.get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_carga_masiva.xlsx"}
    )

@app.post("/api/excel/importar-unificado")
async def importar_unificado(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Por favor sube un archivo Excel válido (.xlsx)")
        
    try:
        contents = await file.read()
        file_stream = io.BytesIO(contents)
        wb = openpyxl.load_workbook(file_stream, data_only=True)
        
        # 0. PROCESAR HOJA DE UBICACIONES
        locations_imported = 0
        locations_updated = 0
        
        if "Importar Ubicaciones" in wb.sheetnames:
            sheet_ubicaciones = wb["Importar Ubicaciones"]
            rows_ubicaciones = list(sheet_ubicaciones.iter_rows())
            
            for idx, row in enumerate(rows_ubicaciones[1:], start=2):
                if not row or not any(cell.value for cell in row):
                    continue
                
                planta_name = str(row[0].value).strip() if row[0].value is not None else ""
                edificio_name = str(row[1].value).strip() if row[1].value is not None else ""
                ubicacion_name = str(row[2].value).strip() if row[2].value is not None else ""
                codigo_val = str(row[3].value).strip() if row[3].value is not None else ""
                uso_val = str(row[4].value).strip() if row[4].value is not None else "Oficina"
                cargo_val = str(row[5].value).strip() if row[5].value is not None else ""
                ocupantes_val = str(row[6].value).strip() if row[6].value is not None else ""
                
                if not planta_name or not edificio_name or not ubicacion_name:
                    continue
                
                # Detect color in any cell of this row
                row_color = None
                for cell in row:
                    c_color = serialize_cell_color(cell)
                    if c_color:
                        row_color = c_color
                        break
                
                # Buscar o crear Planta
                planta = db.exec(select(Planta).where(Planta.nombre == planta_name)).first()
                if not planta:
                    planta = Planta(nombre=planta_name)
                    db.add(planta)
                    db.commit()
                    db.refresh(planta)
                
                # Buscar o crear Edificio
                edificio = db.exec(
                    select(Edificio)
                    .where(Edificio.nombre == edificio_name, Edificio.planta_id == planta.id)
                ).first()
                if not edificio:
                    edificio = Edificio(nombre=edificio_name, planta_id=planta.id)
                    db.add(edificio)
                    db.commit()
                    db.refresh(edificio)
                
                # Buscar Ubicación por código (si está presente) o por nombre
                ubicacion = None
                if codigo_val:
                    ubicacion = db.exec(
                        select(Ubicacion)
                        .where(Ubicacion.codigo == codigo_val)
                        .where(Ubicacion.edificio_id == edificio.id)
                    ).first()
                
                if not ubicacion:
                    ubicacion = db.exec(
                        select(Ubicacion)
                        .where(Ubicacion.nombre == ubicacion_name)
                        .where(Ubicacion.edificio_id == edificio.id)
                    ).first()
                
                if ubicacion:
                    # Actualizar ubicación
                    ubicacion.nombre = ubicacion_name
                    ubicacion.codigo = codigo_val if codigo_val else None
                    ubicacion.uso = uso_val if uso_val else "Oficina"
                    ubicacion.cargo = cargo_val if cargo_val else None
                    ubicacion.color = row_color
                    db.add(ubicacion)
                    db.commit()
                    db.refresh(ubicacion)
                    locations_updated += 1
                else:
                    # Crear nueva ubicación
                    ubicacion = Ubicacion(
                        nombre=ubicacion_name,
                        codigo=codigo_val if codigo_val else None,
                        uso=uso_val if uso_val else "Oficina",
                        cargo=cargo_val if cargo_val else None,
                        color=row_color,
                        edificio_id=edificio.id
                    )
                    db.add(ubicacion)
                    db.commit()
                    db.refresh(ubicacion)
                    locations_imported += 1
                
                # Actualizar ocupantes
                new_occupants = [o.strip() for o in ocupantes_val.split(",") if o.strip()] if ocupantes_val else []
                
                # Limpiar ocupantes existentes
                from sqlalchemy import delete
                db.exec(delete(OcupanteUbicacion).where(OcupanteUbicacion.ubicacion_id == ubicacion.id))
                db.commit()
                
                # Añadir los nuevos
                for o_name in new_occupants:
                    db.add(OcupanteUbicacion(nombre=o_name, ubicacion_id=ubicacion.id))
                
                db.commit()

        # 1. PROCESAR HOJA DE ACTIVOS
        assets_imported = 0
        assets_omitted = 0
        
        sheet_activos = None
        if "Importar Activos" in wb.sheetnames:
            sheet_activos = wb["Importar Activos"]
        elif "Importar Ubicaciones" not in wb.sheetnames:
            sheet_activos = wb.active
            
        rows_activos = []
        if sheet_activos:
            rows_activos = list(sheet_activos.iter_rows())
        
        # Guardar conteo de activos en el Excel para resolver duplicados por cantidad
        seen_in_excel = {}
        valid_activo_rows = []
        
        for idx, row in enumerate(rows_activos[1:], start=2):
            if not row or not any(cell.value for cell in row):
                continue
                
            planta_name = str(row[0].value).strip() if row[0].value is not None else ""
            edificio_name = str(row[1].value).strip() if row[1].value is not None else ""
            ubicacion_name = str(row[2].value).strip() if row[2].value is not None else ""
            activo_name = str(row[3].value).strip() if row[3].value is not None else ""
            tipo_val = str(row[4].value).strip() if row[4].value is not None else "Climatización"
            marca_val = str(row[5].value).strip() if row[5].value is not None else None
            modelo_val = str(row[6].value).strip() if row[6].value is not None else None
            serie_val = str(row[7].value).strip() if row[7].value is not None else None
            
            # Fila de ejemplo (Santa Adela / Fancoil Muro Oficina 1) se omite
            if (planta_name == "Santa Adela" and edificio_name == "Edificio Corporativo" and 
                "Oficina Gerencia" in ubicacion_name and activo_name == "Fancoil Muro Oficina 1"):
                continue
                
            if not planta_name or not edificio_name or not ubicacion_name or not activo_name:
                continue
                
            # Detect color in any cell of this row
            row_color = None
            for cell in row:
                c_color = serialize_cell_color(cell)
                if c_color:
                    row_color = c_color
                    break
                    
            valid_activo_rows.append({
                "idx": idx,
                "planta": planta_name,
                "edificio": edificio_name,
                "ubicacion": ubicacion_name,
                "nombre": activo_name,
                "tipo": tipo_val,
                "marca": marca_val,
                "modelo": modelo_val,
                "serie": serie_val,
                "color": row_color
            })
            
            # Contabilizar apariciones en Excel
            key = (planta_name, edificio_name, ubicacion_name, activo_name, serie_val)
            seen_in_excel[key] = seen_in_excel.get(key, 0) + 1
            
        # Procesar los activos
        processed_keys = set()
        for item in valid_activo_rows:
            key = (item["planta"], item["edificio"], item["ubicacion"], item["nombre"], item["serie"])
            if key in processed_keys:
                continue
            processed_keys.add(key)
            
            # Resolver jerarquía de ubicación
            planta = db.exec(select(Planta).where(Planta.nombre == item["planta"])).first()
            if not planta:
                planta = Planta(nombre=item["planta"])
                db.add(planta)
                db.commit()
                db.refresh(planta)
                
            edificio = db.exec(select(Edificio).where(Edificio.nombre == item["edificio"], Edificio.planta_id == planta.id)).first()
            if not edificio:
                edificio = Edificio(nombre=item["edificio"], planta_id=planta.id)
                db.add(edificio)
                db.commit()
                db.refresh(edificio)
                
            ubicacion = db.exec(select(Ubicacion).where(Ubicacion.nombre == item["ubicacion"], Ubicacion.edificio_id == edificio.id)).first()
            if not ubicacion:
                # Buscar coincidencia parcial (ej. código)
                statement = select(Ubicacion).where(Ubicacion.edificio_id == edificio.id)
                all_ubics = db.exec(statement).all()
                matched_u = None
                for u in all_ubics:
                    if u.nombre.startswith(item["ubicacion"]):
                        matched_u = u
                        break
                if matched_u:
                    ubicacion = matched_u
                else:
                    ubicacion = Ubicacion(nombre=item["ubicacion"], edificio_id=edificio.id)
                    db.add(ubicacion)
                    db.commit()
                    db.refresh(ubicacion)
                    
            # Conteo en la base de datos
            db_assets = []
            if item["serie"]:
                db_assets = db.exec(
                    select(Activo)
                    .where(Activo.numero_serie == item["serie"])
                    .where(Activo.ubicacion_id == ubicacion.id)
                    .where(Activo.estado != "Reemplazado")
                    .where(Activo.estado != "Eliminado sin Reemplazo")
                ).all()
            else:
                db_assets = db.exec(
                    select(Activo)
                    .where(Activo.nombre == item["nombre"])
                    .where(Activo.ubicacion_id == ubicacion.id)
                    .where(Activo.estado != "Reemplazado")
                    .where(Activo.estado != "Eliminado sin Reemplazo")
                ).all()
            db_count = len(db_assets)
            
            # Update existing assets colors if imported
            for a in db_assets:
                a.color = item["color"]
                db.add(a)
                
            excel_count = seen_in_excel[key]
            
            # Crear la diferencia
            to_import = excel_count - db_count
            if to_import > 0:
                for _ in range(to_import):
                    new_activo = Activo(
                        nombre=item["nombre"],
                        tipo=item["tipo"],
                        marca=item["marca"],
                        modelo=item["modelo"],
                        numero_serie=item["serie"],
                        estado="Operativo",
                        ubicacion_id=ubicacion.id,
                        color=item["color"]
                    )
                    db.add(new_activo)
                assets_imported += to_import
            else:
                assets_omitted += excel_count
                
        db.commit()
        
        # 2. PROCESAR HOJA DE DESPIECES
        comps_imported = 0
        comps_updated = 0
        comps_omitted = 0
        unmatched_comps = []
        
        if "Importar Despieces" in wb.sheetnames:
            sheet_despieces = wb["Importar Despieces"]
            rows_despieces = list(sheet_despieces.iter_rows(values_only=True))
            
            for idx, row in enumerate(rows_despieces[1:], start=2):
                if not row or not any(row):
                    continue
                    
                # Leer columnas
                id_comp_str = str(row[0]).strip() if row[0] is not None else ""
                serie_activo = str(row[1]).strip() if row[1] is not None else ""
                planta_name = str(row[2]).strip() if row[2] is not None else ""
                edificio_name = str(row[3]).strip() if row[3] is not None else ""
                ubicacion_name = str(row[4]).strip() if row[4] is not None else ""
                activo_name = str(row[5]).strip() if row[5] is not None else ""
                comp_name = str(row[6]).strip() if row[6] is not None else ""
                marca_val = str(row[7]).strip() if row[7] is not None else None
                modelo_val = str(row[8]).strip() if row[8] is not None else None
                comp_serie = str(row[9]).strip() if row[9] is not None else None
                estado_val = str(row[10]).strip() if row[10] is not None else "Operativo"
                
                # Fila de ejemplo (ejemplo Compresor Carrier 42GW o PZ-1) se omite
                if comp_name == "Compresor" and (serie_activo == "FC-P1A-E1" or id_comp_str == "PZ-1"):
                    continue
                    
                if not comp_name:
                    continue
                    
                # Limpiar ID si se provee
                comp_id = None
                if id_comp_str:
                    clean_id = id_comp_str.upper().replace("PZ-", "").replace("PZ", "").strip()
                    if clean_id.isdigit():
                        comp_id = int(clean_id)
                        
                # Si se provee ID, intentar actualizar
                existing_comp = None
                if comp_id:
                    existing_comp = db.get(ComponenteActivo, comp_id)
                    
                if existing_comp:
                    # Buscar y asociar activo
                    activo = None
                    if serie_activo:
                        activo = db.exec(
                            select(Activo)
                            .where(Activo.numero_serie == serie_activo)
                            .where(Activo.estado != "Reemplazado")
                            .where(Activo.estado != "Eliminado sin Reemplazo")
                        ).first()
                        
                    if not activo and activo_name and planta_name and edificio_name and ubicacion_name:
                        planta = db.exec(select(Planta).where(Planta.nombre == planta_name)).first()
                        if planta:
                            edificio = db.exec(select(Edificio).where(Edificio.nombre == edificio_name, Edificio.planta_id == planta.id)).first()
                            if edificio:
                                ubicacion = db.exec(select(Ubicacion).where(Ubicacion.nombre == ubicacion_name, Ubicacion.edificio_id == edificio.id)).first()
                                if ubicacion:
                                    activo = db.exec(
                                        select(Activo)
                                        .where(Activo.nombre == activo_name)
                                        .where(Activo.ubicacion_id == ubicacion.id)
                                        .where(Activo.estado != "Reemplazado")
                                        .where(Activo.estado != "Eliminado sin Reemplazo")
                                    ).first()
                    
                    # Actualizar datos
                    existing_comp.nombre = comp_name
                    existing_comp.marca = marca_val
                    existing_comp.modelo = modelo_val
                    existing_comp.numero_serie = comp_serie
                    existing_comp.estado = estado_val
                    if activo:
                        existing_comp.activo_id = activo.id
                        
                    db.add(existing_comp)
                    comps_updated += 1
                    continue
                    
                # Si se especificó un ID pero no se encontró en base de datos, lo tratamos como nuevo
                if comp_id and not existing_comp:
                    pass
                    
                # Buscar activo para nuevo componente
                activo = None
                if serie_activo:
                    activo = db.exec(
                        select(Activo)
                        .where(Activo.numero_serie == serie_activo)
                        .where(Activo.estado != "Reemplazado")
                        .where(Activo.estado != "Eliminado sin Reemplazo")
                    ).first()
                    
                if not activo and activo_name and planta_name and edificio_name and ubicacion_name:
                    planta = db.exec(select(Planta).where(Planta.nombre == planta_name)).first()
                    if planta:
                        edificio = db.exec(select(Edificio).where(Edificio.nombre == edificio_name, Edificio.planta_id == planta.id)).first()
                        if edificio:
                            ubicacion = db.exec(select(Ubicacion).where(Ubicacion.nombre == ubicacion_name, Ubicacion.edificio_id == edificio.id)).first()
                            if ubicacion:
                                activo = db.exec(
                                    select(Activo)
                                    .where(Activo.nombre == activo_name)
                                    .where(Activo.ubicacion_id == ubicacion.id)
                                    .where(Activo.estado != "Reemplazado")
                                    .where(Activo.estado != "Eliminado sin Reemplazo")
                                ).first()
                                
                if not activo:
                    unmatched_comps.append(f"Fila {idx}: no se encontró activo para Serie '{serie_activo}' o Nombre '{activo_name}'")
                    continue
                    
                # Verificar duplicado de componente en el mismo activo
                existing_comps = db.exec(
                    select(ComponenteActivo)
                    .where(ComponenteActivo.activo_id == activo.id)
                    .where(ComponenteActivo.nombre == comp_name)
                ).all()
                
                is_duplicate = False
                for ec in existing_comps:
                    if ec.modelo == modelo_val and ec.numero_serie == comp_serie:
                        is_duplicate = True
                        break
                        
                if is_duplicate:
                    comps_omitted += 1
                    continue
                    
                # Insertar componente
                new_comp = ComponenteActivo(
                    nombre=comp_name,
                    marca=marca_val,
                    modelo=modelo_val,
                    numero_serie=comp_serie,
                    estado=estado_val,
                    activo_id=activo.id
                )
                db.add(new_comp)
                comps_imported += 1
                
            db.commit()
            
        # Mensaje de resultado
        parts = []
        if "Importar Ubicaciones" in wb.sheetnames:
            parts.append(f"{locations_imported + locations_updated} ubicaciones ({locations_imported} creadas, {locations_updated} actualizadas)")
        parts.append(f"{assets_imported} activos creados ({assets_omitted} omitidos)")
        if "Importar Despieces" in wb.sheetnames:
            parts.append(f"{comps_imported + comps_updated} componentes ({comps_imported} creados, {comps_updated} actualizados, {comps_omitted} omitidos)")
            
        msg = "Se importaron con éxito: " + ", ".join(parts) + "."
        if unmatched_comps:
            msg += f" Omitidos {len(unmatched_comps)} por activo o ID no encontrado: {', '.join(unmatched_comps[:2])}..."
            
        return {
            "status": "success",
            "locations_imported": locations_imported,
            "locations_updated": locations_updated,
            "assets_imported": assets_imported,
            "assets_omitted": assets_omitted,
            "comps_imported": comps_imported,
            "comps_updated": comps_updated,
            "comps_omitted": comps_omitted,
            "message": msg
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al procesar archivo Excel: {str(e)}")

# Montar los archivos estáticos para la interfaz de usuario
app.mount("/static", StaticFiles(directory="static"), name="static")

import uvicorn
import os

if __name__ == "__main__":
    host = os.getenv("HOST", "0.0.0.0" if os.getenv("RENDER") else "127.0.0.1")
    port = int(os.getenv("PORT", 8000))
    reload_mode = not os.getenv("RENDER")
    uvicorn.run("main:app", host=host, port=port, reload=reload_mode)
