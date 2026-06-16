import os
import re
from sqlmodel import SQLModel, create_engine, Session, select
import openpyxl

from models import (
    Planta, Edificio, Ubicacion, Activo, ComponenteActivo,
    Tecnico, OrdenTrabajo, PlantillaChequeo, ItemPlantillaChequeo,
    RespuestaChequeo, FotoOrdenTrabajo, ComentarioAvanceOT, OcupanteUbicacion,
    OrdenTrabajoComponente
)

# Render uses DATABASE_URL for PostgreSQL connection
database_url = os.getenv("DATABASE_URL", "sqlite:///fms.db")
if database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)

connect_args = {}
if database_url.startswith("sqlite"):
    connect_args = {"check_same_thread": False}

engine = create_engine(database_url, connect_args=connect_args)

def create_db_and_tables():
    SQLModel.metadata.create_all(engine)
    try:
        from sqlalchemy import inspect, text
        inspector = inspect(engine)
        columns = [col['name'] for col in inspector.get_columns('ordentrabajo')]
        if 'fecha_programada' not in columns:
            print("Migrando base de datos: agregando columna 'fecha_programada' a la tabla 'ordentrabajo'...")
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE ordentrabajo ADD COLUMN fecha_programada TIMESTAMP"))
            print("Columna 'fecha_programada' agregada exitosamente.")
        if 'fecha_inicio' not in columns:
            print("Migrando base de datos: agregando columna 'fecha_inicio' a la tabla 'ordentrabajo'...")
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE ordentrabajo ADD COLUMN fecha_inicio TIMESTAMP"))
            print("Columna 'fecha_inicio' agregada exitosamente.")
        if 'estado_ejecucion' not in columns:
            print("Migrando base de datos: agregando columna 'estado_ejecucion' a la tabla 'ordentrabajo'...")
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE ordentrabajo ADD COLUMN estado_ejecucion VARCHAR(50) DEFAULT 'NO_INICIADA'"))
                conn.execute(text("UPDATE ordentrabajo SET estado_ejecucion = 'REALIZADA' WHERE fecha_resolucion IS NOT NULL"))
                conn.execute(text("UPDATE ordentrabajo SET estado_ejecucion = 'EN_PROCESO' WHERE fecha_inicio IS NOT NULL AND fecha_resolucion IS NULL"))
            print("Columna 'estado_ejecucion' agregada exitosamente.")

        # Check and migrate plantillachequeo table
        columns_pc = [col['name'] for col in inspector.get_columns('plantillachequeo')]
        if 'tipo_revision' not in columns_pc:
            print("Migrando base de datos: agregando columna 'tipo_revision' a la tabla 'plantillachequeo'...")
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE plantillachequeo ADD COLUMN tipo_revision VARCHAR(100) DEFAULT 'Chequeo Preventivo'"))
            print("Columna 'tipo_revision' agregada exitosamente.")
        
        # Ensure existing templates are categorized correctly
        with engine.begin() as conn:
            conn.execute(text("UPDATE plantillachequeo SET tipo_revision = 'Ronda' WHERE nombre LIKE '%Ronda%'"))

        # Check and migrate ubicacion table
        columns_ub = [col['name'] for col in inspector.get_columns('ubicacion')]
        if 'codigo' not in columns_ub:
            print("Migrando base de datos: agregando columna 'codigo' a la tabla 'ubicacion'...")
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE ubicacion ADD COLUMN codigo VARCHAR(100)"))
            print("Columna 'codigo' agregada exitosamente.")
        if 'uso' not in columns_ub:
            print("Migrando base de datos: agregando columna 'uso' a la tabla 'ubicacion'...")
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE ubicacion ADD COLUMN uso VARCHAR(100) DEFAULT 'Oficina'"))
            print("Columna 'uso' agregada exitosamente.")
        if 'cargo' not in columns_ub:
            print("Migrando base de datos: agregando columna 'cargo' a la tabla 'ubicacion'...")
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE ubicacion ADD COLUMN cargo VARCHAR(200)"))
            print("Columna 'cargo' agregada exitosamente.")
        if 'color' not in columns_ub:
            print("Migrando base de datos: agregando columna 'color' a la tabla 'ubicacion'...")
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE ubicacion ADD COLUMN color VARCHAR(50)"))
            print("Columna 'color' agregada exitosamente.")

        # Check and migrate activo table
        columns_act = [col['name'] for col in inspector.get_columns('activo')]
        if 'color' not in columns_act:
            print("Migrando base de datos: agregando columna 'color' a la tabla 'activo'...")
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE activo ADD COLUMN color VARCHAR(50)"))
            print("Columna 'color' agregada exitosamente.")

        # Execute data migration for existing 129 locations
        try:
            with Session(engine) as session:
                stmt = select(Ubicacion).where(Ubicacion.codigo == None)
                unmigrated = session.exec(stmt).all()
                if unmigrated:
                    print(f"Migrando {len(unmigrated)} registros de ubicaciones en la base de datos...")
                    import re
                    
                    # Palabras clave genéricas para identificar el uso (en minúsculas)
                    usos_genericos = {
                        "baño": "Baño", "wc": "Baño", "sanitario": "Baño", "camarin": "Camarín", "camarines": "Camarín",
                        "bodega": "Bodega", "cocina": "Cocina", "kitchinet": "Kitchenette", "kitchenette": "Kitchenette",
                        "servidores": "Servidores", "reuniones": "Sala de Reuniones", "directorio": "Sala de Reuniones",
                        "comedor": "Comedor", "casino": "Casino", "pasillo": "Pasillo", "hall": "Hall", "recepcion": "Recepción",
                        "patio": "Patio", "terraza": "Terraza", "estacionamiento": "Estacionamiento", "taller": "Taller",
                        "ducto": "Ducto", "horno": "Horno"
                    }
                    
                    # Palabras clave de roles/cargos/departamentos (en minúsculas)
                    roles_cargos = [
                        "jefe", "gte", "gerente", "director", "asistente", "asit", "gn", "auditor", "supervisor",
                        "subgerente", "coordinador", "construccion", "construcción", "rental", "marketing", "marketin",
                        "izaje", "ti", "finanzas", "operaciones", "personas", "rrhh", "tesoreria", "tesorería"
                    ]
                    
                    for u in unmigrated:
                        # Extraer el código entre corchetes, ej: "Oficina RRHH [EC-P1-RecN-#1]"
                        match_code = re.search(r'\[([^\]]+)\]', u.nombre)
                        if match_code:
                            code_val = match_code.group(1).strip()
                            # Extraer la descripción antes de los corchetes
                            desc_val = re.sub(r'\[[^\]]+\]', '', u.nombre).strip()
                            
                            u.codigo = code_val
                            
                            # Analizar la descripción para determinar uso, cargo y ocupantes
                            desc_lower = desc_val.lower()
                            
                            # 1. Comprobar si es un uso genérico
                            matched_uso = None
                            for key, val in usos_genericos.items():
                                if key in desc_lower:
                                    matched_uso = val
                                    break
                            
                            if matched_uso:
                                u.uso = matched_uso
                                u.cargo = None
                                u.nombre = f"{matched_uso} [{code_val}]"
                            else:
                                # 2. Comprobar si es un cargo/rol/departamento
                                is_cargo = any(role in desc_lower for role in roles_cargos)
                                if is_cargo:
                                    u.uso = "Oficina"
                                    u.cargo = desc_val
                                    u.nombre = f"{desc_val} [{code_val}]"
                                else:
                                    # 3. Asumir que es una persona/ocupante (si no está vacío y no es "oficina compartida")
                                    if desc_val and desc_val.lower() != "oficina compartida":
                                        u.uso = "Oficina"
                                        u.cargo = None
                                        u.nombre = f"Oficina [{code_val}]"
                                        # Agregar ocupante
                                        new_ocupante = OcupanteUbicacion(nombre=desc_val, ubicacion_id=u.id)
                                        session.add(new_ocupante)
                                    else:
                                        # Oficina compartida o sin nombre
                                        u.uso = "Oficina"
                                        u.cargo = None
                                        u.nombre = f"Oficina [{code_val}]"
                            
                            session.add(u)
                    session.commit()
                    print("Migración de ubicaciones completada con éxito.")
        except Exception as ex_mig:
            print(f"Error durante la migración de datos de ubicaciones: {ex_mig}")

        # Repair any misclassified bathrooms (e.g. Gte General [EC-P2-OS-bñ-6])
        try:
            with Session(engine) as session:
                misclassified_bathrooms = session.exec(
                    select(Ubicacion)
                    .where(Ubicacion.uso == "Oficina")
                    .where(Ubicacion.nombre.like("%bñ%") | Ubicacion.codigo.like("%bñ%") | Ubicacion.nombre.like("%bn%") | Ubicacion.codigo.like("%bn%"))
                ).all()
                if misclassified_bathrooms:
                    print(f"Reparando {len(misclassified_bathrooms)} baños mal clasificados...")
                    for u in misclassified_bathrooms:
                        u.uso = "Baño"
                        if "Gte" in u.nombre:
                            u.cargo = "Gte General"
                        if u.codigo:
                            u.nombre = f"Baño [{u.codigo}]"
                        session.add(u)
                    session.commit()
                    print("Baños reparados con éxito.")
        except Exception as ex_rep:
            print(f"Error al reparar baños: {ex_rep}")

    except Exception as e:
        print(f"Error al verificar/migrar columnas de base de datos: {e}")


def normalize_code(code):
    if not code:
        return ""
    # Convert to uppercase and remove spaces
    c = str(code).upper().replace(" ", "").strip()
    # Normalize zero-padded numbers, e.g., OF01 -> OF1, OF-01 -> OF-1
    c = re.sub(r'([A-Z]+)0+([1-9]+)', r'\1\2', c)
    # Also handle trailing dashes or hashes
    c = c.replace("#", "")
    return c

def seed_database():
    with Session(engine) as session:
        # Migrate any existing full names in DB to initials in-place
        name_map = {
            "Javier Pinochet": "JP",
            "Alex Valenzuela": "AV",
            "Simón Monardes": "SM",
            "Simon Monardes": "SM",
            "Víctor Hugo Hurtado": "VHH",
            "Victor Hugo Hurtado": "VHH",
            "Víctor Parra": "VP",
            "Victor Parra": "VP"
        }
        
        def get_initials(name):
            if not name:
                return name
            if name in name_map:
                return name_map[name]
            if len(name) <= 3 and name.isupper():
                return name
            parts = [p for p in name.split() if p]
            if not parts:
                return name
            return "".join([p[0].upper() for p in parts])

        # 1. Update Tecnico table
        techs_in_db = session.exec(select(Tecnico)).all()
        for t in techs_in_db:
            old_name = t.nombre
            new_name = get_initials(old_name)
            if old_name != new_name:
                t.nombre = new_name
                session.add(t)
                print(f"- Migrando Tecnico ID {t.id}: '{old_name}' -> '{new_name}'")
                
        # 2. Update OrdenTrabajo table
        ots = session.exec(select(OrdenTrabajo)).all()
        for ot in ots:
            modified = False
            if ot.reportado_por:
                orig = ot.reportado_por
                for full, init in name_map.items():
                    if full in orig:
                        orig = orig.replace(full, init)
                        modified = True
                if modified:
                    ot.reportado_por = orig
                    
            if ot.comentarios_tecnicos:
                orig_comments = ot.comentarios_tecnicos
                temp_comments = orig_comments
                for full, init in name_map.items():
                    if full in temp_comments:
                        temp_comments = temp_comments.replace(full, init)
                if temp_comments != orig_comments:
                    ot.comentarios_tecnicos = temp_comments
                    modified = True
                    
            if modified:
                session.add(ot)

        # 3. Update ComentarioAvanceOT table
        comments = session.exec(select(ComentarioAvanceOT)).all()
        for c in comments:
            modified = False
            if c.autor:
                orig_autor = c.autor
                new_autor = get_initials(orig_autor)
                if orig_autor != new_autor:
                    c.autor = new_autor
                    modified = True
                    
            if c.comentario:
                orig_txt = c.comentario
                temp_txt = orig_txt
                for full, init in name_map.items():
                    if full in temp_txt:
                        temp_txt = temp_txt.replace(full, init)
                if temp_txt != orig_txt:
                    c.comentario = temp_txt
                    modified = True
                    
            if modified:
                session.add(c)
                
        session.commit()

        # Check and update technicians if they do not match the new requested list
        new_tech_names = [
            "JP",
            "AV",
            "SM",
            "VHH",
            "VP"
        ]
        try:
            techs_in_db = session.exec(select(Tecnico).order_by(Tecnico.id)).all()
            tech_names_in_db = [t.nombre for t in techs_in_db]
            
            if tech_names_in_db != new_tech_names:
                print("Actualizando lista de técnicos en la base de datos...")
                # Clear references to avoid dangling foreign keys
                ots = session.exec(select(OrdenTrabajo)).all()
                for ot in ots:
                    if ot.tecnico_id:
                        t = session.get(Tecnico, ot.tecnico_id)
                        if not t or t.nombre not in new_tech_names:
                            ot.tecnico_id = None
                            if ot.estado == "En Proceso":
                                ot.estado = "Pendiente"
                            session.add(ot)
                session.commit()

                # Delete old techs
                for t in techs_in_db:
                    session.delete(t)
                session.commit()

                # Insert new techs
                t1 = Tecnico(nombre="JP", email="javier.pinochet@emaresa.cl", especialidad="Climatización")
                t2 = Tecnico(nombre="AV", email="alex.valenzuela@emaresa.cl", especialidad="Climatización")
                t3 = Tecnico(nombre="SM", email="simon.monardes@emaresa.cl", especialidad="Climatización")
                t4 = Tecnico(nombre="VHH", email="victor.hurtado@emaresa.cl", especialidad="Climatización")
                t5 = Tecnico(nombre="VP", email="victor.parra@emaresa.cl", especialidad="Climatización")
                session.add_all([t1, t2, t3, t4, t5])
                session.commit()
                print("Lista de técnicos actualizada exitosamente.")
        except Exception as e:
            print(f"Error al verificar/actualizar técnicos: {e}")
            session.rollback()

        # Check if database is already seeded
        statement = select(Planta)
        existing_plants = session.exec(statement).all()
        if existing_plants:
            print("Base de datos ya inicializada con plantas.")
            return

        print("Iniciando carga de datos de plantas...")
        # Create Plants
        p1 = Planta(nombre="Santa Adela")
        p2 = Planta(nombre="La Divisa")
        p3 = Planta(nombre="Sta. A. 10.000")
        
        session.add_all([p1, p2, p3])
        session.commit()
        
        # Refresh to get IDs
        session.refresh(p1)
        session.refresh(p2)
        session.refresh(p3)
        
        # Buildings for Santa Adela matching Excel ZONA
        b_santa_adela = [
            "Edificio Corporativo", "Galpón 1", "Galpón 2", 
            "Galpón 3", "Galpón 4", "Galpón 5", "Galpón 6", 
            "Centro de Distriducion", "Patio las torres"
        ]
        edificios_sta_adela = {}
        for name in b_santa_adela:
            ed = Edificio(nombre=name, planta_id=p1.id)
            session.add(ed)
            session.commit()
            session.refresh(ed)
            edificios_sta_adela[name] = ed.id
            
        # Buildings for La Divisa
        b_la_divisa = [
            "Edificio Corporativo", "Nave 0", "Nave 1", 
            "Nave 2", "Nave 3", "Camarines", "Casino"
        ]
        for name in b_la_divisa:
            session.add(Edificio(nombre=name, planta_id=p2.id))
            
        # Buildings for Sta. A. 10.000
        b_sta_a = [
            "Galpón 1", "Galpón 2", "Galpón 3", 
            "Galpón 4A", "Galpón 4B", "Galpón 5", 
            "Galpón 6", "Galpón 7", "Comedor", 
            "Camarines P1", "Horno secado"
        ]
        for name in b_sta_a:
            session.add(Edificio(nombre=name, planta_id=p3.id))
            
        # Add some initial Technicians
        t1 = Tecnico(nombre="JP", email="javier.pinochet@emaresa.cl", especialidad="Climatización")
        t2 = Tecnico(nombre="AV", email="alex.valenzuela@emaresa.cl", especialidad="Climatización")
        t3 = Tecnico(nombre="SM", email="simon.monardes@emaresa.cl", especialidad="Climatización")
        t4 = Tecnico(nombre="VHH", email="victor.hurtado@emaresa.cl", especialidad="Climatización")
        t5 = Tecnico(nombre="VP", email="victor.parra@emaresa.cl", especialidad="Climatización")
        session.add_all([t1, t2, t3, t4, t5])
        session.commit()

        # Dictionary to keep track of locations by normalized code
        loc_code_to_id = {}

        # Load Locations for Santa Adela from Excel
        excel_zones_path = r"BD para ubicaciones/Codificacion zonas 25 07 01 945.xlsx"
        if os.path.exists(excel_zones_path):
            print("Cargando ubicaciones de Santa Adela desde Excel...")
            wb = openpyxl.load_workbook(excel_zones_path, read_only=True)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                zona = row[0]       # ZONA (Edificio)
                ubic_name = row[4]  # Ubicacion (Nombre)
                code = row[5]       # COD ubicacion
                
                if not zona or not ubic_name or not code:
                    continue
                
                zona = str(zona).strip()
                ubic_name = str(ubic_name).strip()
                code = str(code).strip()
                
                # Find building ID
                edificio_id = edificios_sta_adela.get(zona)
                if not edificio_id:
                    new_ed = Edificio(nombre=zona, planta_id=p1.id)
                    session.add(new_ed)
                    session.commit()
                    session.refresh(new_ed)
                    edificios_sta_adela[zona] = new_ed.id
                    edificio_id = new_ed.id
                
                loc_full_name = f"{ubic_name} [{code}]"
                new_loc = Ubicacion(nombre=loc_full_name, edificio_id=edificio_id)
                session.add(new_loc)
                session.commit()
                session.refresh(new_loc)
                
                norm_code = normalize_code(code)
                loc_code_to_id[norm_code] = new_loc.id
            print(f"Se cargaron {len(loc_code_to_id)} ubicaciones en Santa Adela.")

        # Load Fancoils and Despieces from Excel
        excel_fc_path = r"BD para ubicaciones/levantamiento Fancoil 25 12 19 1125.xlsx"
        if os.path.exists(excel_fc_path):
            print("Cargando Fancoils y despieces de Santa Adela...")
            wb = openpyxl.load_workbook(excel_fc_path, read_only=True)
            sheet = wb.active
            
            # Group rows by equipment (CODIGO AA)
            equip_rows = {}
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                eq_code = row[3]  # CODIGO AA
                if not eq_code:
                    continue
                eq_code = str(eq_code).strip()
                if eq_code not in equip_rows:
                    equip_rows[eq_code] = []
                equip_rows[eq_code].append(row)
                
            # Create each equipment
            assets_created = 0
            for eq_code, rows in equip_rows.items():
                main_row = None
                for r in rows:
                    pieza_name = str(r[7]).strip().lower() if r[7] else ""
                    if "fancoil" in pieza_name or "split" in pieza_name:
                        main_row = r
                        break
                if not main_row:
                    main_row = rows[0]
                
                # Get location
                loc_code_raw = main_row[2]
                loc_name_raw = main_row[1]
                
                loc_id = None
                if loc_code_raw:
                    norm_loc_code = normalize_code(loc_code_raw)
                    loc_id = loc_code_to_id.get(norm_loc_code)
                
                # Dynamic location creation if missing
                if not loc_id and loc_name_raw:
                    loc_name_clean = str(loc_name_raw).strip()
                    loc_code_clean = str(loc_code_raw).strip() if loc_code_raw else "S/C"
                    loc_full_name = f"{loc_name_clean} [{loc_code_clean}]"
                    
                    # Target 'Edificio Corporativo' as default for unmatched fancoils
                    ed_id = edificios_sta_adela.get("Edificio Corporativo")
                    if not ed_id:
                        # Fetch the first Santa Adela building if Edificio Corporativo is not loaded
                        ed_id = list(edificios_sta_adela.values())[0]
                        
                    new_loc = Ubicacion(nombre=loc_full_name, edificio_id=ed_id)
                    session.add(new_loc)
                    session.commit()
                    session.refresh(new_loc)
                    loc_id = new_loc.id
                    if loc_code_raw:
                        loc_code_to_id[normalize_code(loc_code_raw)] = loc_id
                
                # Check for split or fancoil type
                is_split = "split" in str(main_row[7]).lower() if main_row[7] else False
                prefix = "Aire Acondicionado Split" if is_split else "Fancoil"
                asset_name = f"{prefix} {eq_code}"
                model_val = str(main_row[8]).strip() if main_row[8] else None
                
                new_asset = Activo(
                    nombre=asset_name,
                    tipo="Climatizacion",
                    marca="Carrier" if not is_split else "S/M",
                    modelo=model_val,
                    numero_serie=eq_code,
                    estado="Operativo",
                    ubicacion_id=loc_id
                )
                session.add(new_asset)
                session.commit()
                session.refresh(new_asset)
                assets_created += 1
                
                # Load despiece components
                for r in rows:
                    pieza_name = r[7]
                    if not pieza_name:
                        continue
                    pieza_name_clean = str(pieza_name).strip()
                    comp_model = str(r[8]).strip() if r[8] else None
                    comp_obs = str(r[9]).strip() if r[9] else ""
                    
                    new_comp = ComponenteActivo(
                        nombre=pieza_name_clean,
                        modelo=comp_model,
                        estado="Operativo",
                        activo_id=new_asset.id,
                        numero_serie=f"Obs: {comp_obs}"[:100] if comp_obs else None
                    )
                    session.add(new_comp)
            session.commit()
            print(f"Se cargaron {assets_created} equipos Fancoil con sus respectivos componentes.")

        # Seed Checklist Templates
        print("Cargando plantillas de chequeo...")
        t_split = PlantillaChequeo(
            nombre="Mantenimiento Preventivo de Aire Acondicionado Split",
            descripcion="Inspección de rutina y mantenimiento preventivo para unidades Split de aire acondicionado.",
            tipo_revision="Chequeo Preventivo"
        )
        session.add(t_split)
        session.commit()
        session.refresh(t_split)
        
        split_items = [
            ItemPlantillaChequeo(texto_pregunta="Limpieza de filtros de aire de unidad evaporadora", tipo_respuesta="booleano", plantilla_id=t_split.id),
            ItemPlantillaChequeo(texto_pregunta="Verificación de drenaje y limpieza de bandeja de condensado", tipo_respuesta="booleano", plantilla_id=t_split.id),
            ItemPlantillaChequeo(texto_pregunta="Medición de corriente de consumo del compresor (Amperios)", tipo_respuesta="numerico", unidad_medida="A", plantilla_id=t_split.id),
            ItemPlantillaChequeo(texto_pregunta="Presión de trabajo de línea de succión (PSI)", tipo_respuesta="numerico", unidad_medida="psi", plantilla_id=t_split.id),
            ItemPlantillaChequeo(texto_pregunta="Limpieza de serpentín condensador (Unidad Exterior)", tipo_respuesta="booleano", plantilla_id=t_split.id),
            ItemPlantillaChequeo(texto_pregunta="Inspección de aislación térmica de tuberías de refrigeración", tipo_respuesta="booleano", plantilla_id=t_split.id),
            ItemPlantillaChequeo(texto_pregunta="Observaciones adicionales de la inspección", tipo_respuesta="texto", plantilla_id=t_split.id)
        ]
        session.add_all(split_items)
        
        t_gen = PlantillaChequeo(
            nombre="Mantenimiento Preventivo de Generador Eléctrico",
            descripcion="Protocolo mensual de inspección y puesta en marcha para generadores de respaldo.",
            tipo_revision="Chequeo Preventivo"
        )
        session.add(t_gen)
        session.commit()
        session.refresh(t_gen)
        
        gen_items = [
            ItemPlantillaChequeo(texto_pregunta="Medición de voltaje de batería en vacío (VCC)", tipo_respuesta="numerico", unidad_medida="V", plantilla_id=t_gen.id),
            ItemPlantillaChequeo(texto_pregunta="Verificación de nivel de aceite de motor y estado general", tipo_respuesta="booleano", plantilla_id=t_gen.id),
            ItemPlantillaChequeo(texto_pregunta="Verificación de nivel de líquido refrigerante de radiador", tipo_respuesta="booleano", plantilla_id=t_gen.id),
            ItemPlantillaChequeo(texto_pregunta="Prueba de arranque y medición de frecuencia en vacío (Hz)", tipo_respuesta="numerico", unidad_medida="Hz", plantilla_id=t_gen.id),
            ItemPlantillaChequeo(texto_pregunta="Nivel de combustible en tanque de almacenamiento (%)", tipo_respuesta="numerico", unidad_medida="%", plantilla_id=t_gen.id),
            ItemPlantillaChequeo(texto_pregunta="Prueba de transferencia y funcionamiento con carga", tipo_respuesta="booleano", plantilla_id=t_gen.id),
            ItemPlantillaChequeo(texto_pregunta="Observaciones técnicas y novedades del generador", tipo_respuesta="texto", plantilla_id=t_gen.id)
        ]
        session.add_all(gen_items)
        
        t_insp = PlantillaChequeo(
            nombre="Ronda de Inspección Semanal (Disperfectos Generales)",
            descripcion="Recorrido de inspección periódica de infraestructura y servicios de la planta.",
            tipo_revision="Ronda"
        )
        session.add(t_insp)
        session.commit()
        session.refresh(t_insp)
        
        insp_items = [
            ItemPlantillaChequeo(texto_pregunta="Luminarias y luces operativas (sin bombillos quemados)", tipo_respuesta="booleano", plantilla_id=t_insp.id),
            ItemPlantillaChequeo(texto_pregunta="Servicios higiénicos y baños conformes (sin filtraciones/grifos malos)", tipo_respuesta="booleano", plantilla_id=t_insp.id),
            ItemPlantillaChequeo(texto_pregunta="Puertas, portones, cerraduras y accesos operativos", tipo_respuesta="booleano", plantilla_id=t_insp.id),
            ItemPlantillaChequeo(texto_pregunta="Orden y limpieza general de pasillos y áreas de trabajo", tipo_respuesta="booleano", plantilla_id=t_insp.id),
            ItemPlantillaChequeo(texto_pregunta="Fugas de fluidos visibles (agua, aire comprimido, refrigeración)", tipo_respuesta="booleano", plantilla_id=t_insp.id),
            ItemPlantillaChequeo(texto_pregunta="Detalles o novedades encontradas en el recorrido", tipo_respuesta="texto", plantilla_id=t_insp.id)
        ]
        session.add_all(insp_items)
        
        session.commit()
        print("Carga de datos iniciales completada con éxito.")

def init_db():
    create_db_and_tables()
    seed_database()
